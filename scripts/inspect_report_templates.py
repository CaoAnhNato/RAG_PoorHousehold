# -*- coding: utf-8 -*-
"""
Module chứa các hàm phục vụ quét, phân tích và trích xuất thông tin cấu trúc,
tiêu đề, địa bàn hành chính (huyện/xã) từ 15 tệp tin biểu mẫu báo cáo Excel (Format_Report).
"""

from __future__ import annotations
import re
from pathlib import Path
from typing import Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from scripts.preprocess_enrich_datasets import clean_text, strip_accents, normalize_key, normalize_no_space
from scripts.export_metadata import build_report_schema_summary

# Danh sách cột dữ liệu đầu vào cơ bản bắt buộc
BASE_INPUT_COLUMNS = [
    "date",
    "family.hostName",
    "family.code",
    "family.coDanTocTaiCho",
    "family.hostGender",
    "family.numberOfMembers",
    "quickReview.result",
    "b1Point",
    "b2Point",
    "classify",
    "reviewer",
]

# Ánh xạ Huyện/Xã dự phòng khi không có tệp biểu mẫu trong thư mục Format_Report
DISTRICT_COMMUNE_FALLBACK = {
    "Huyện Cư Jút": [
        "Thị trấn Ea T'ling",
        "Xã Cư Knia",
        "Xã Đắk DRông",
        "Xã Đắk Wil",
        "Xã Ea Pô",
        "Xã Nam Dong",
        "Xã Tâm Thắng",
        "Xã Trúc Sơn",
    ],
    "Huyện Krông Nô": [
        "Thị trấn Đắk Mâm",
        "Xã Buôn Choáh",
        "Xã Đắk Drô",
        "Xã Đắk N'Drót",
        "Xã Đắk Sôr",
        "Xã Đức Xuyên",
        "Xã Nam Đà",
        "Xã Nam Xuân",
        "Xã Nâm N'Đir",
        "Xã Nâm Nung",
        "Xã Quảng Phú",
        "Xã Tân Thành",
    ],
    "Huyện Tuy Đức": [
        "Xã Đắk Búk So",
        "Xã Đắk Ngo",
        "Xã Đắk R'Tih",
        "Xã Quảng Tâm",
        "Xã Quảng Tân",
        "Xã Quảng Trực",
    ],
    "Huyện Đăk Glong": [
        "Xã Quảng Khê",
        "Xã Quảng Sơn",
        "Xã Quảng Hòa",
        "Xã Đắk Ha",
        "Xã Đắk Som",
        "Xã Đắk R'Măng",
        "Xã Đắk Plao",
    ],
    "Huyện Đắk Mil": [
        "Thị trấn Đắk Mil",
        "Xã Đắk Gằn",
        "Xã Đắk Lao",
        "Xã Đắk N'Drót",
        "Xã Đắk R'La",
        "Xã Đắk Sắk",
        "Xã Đức Mạnh",
        "Xã Đức Minh",
        "Xã Long Sơn",
        "Xã Thuận An",
    ],
    "Huyện Đắk RLấp": [
        "Thị trấn Kiến Đức",
        "Xã Nhân Cơ",
        "Xã Đắk Wer",
        "Xã Kiến Thành",
        "Xã Đạo Nghĩa",
        "Xã Nghĩa Thắng",
        "Xã Nhân Đạo",
        "Xã Đắk Sin",
        "Xã Quảng Tín",
        "Xã Hưng Bình",
        "Xã Đắk Ru",
    ],
    "Huyện Đắk Song": [
        "Thị trấn Đức An",
        "Xã Đắk Hòa",
        "Xã Đắk Môl",
        "Xã Đắk N'Drung",
        "Xã Nam Bình",
        "Xâm N'Jang",
        "Xã Thuận Hà",
        "Xã Thuận Hạnh",
        "Xã Trường Xuân",
    ],
    "Thành phố Gia Nghĩa": [
        "Phường Nghĩa Đức",
        "Phường Nghĩa Thành",
        "Phường Nghĩa Trung",
        "Phường Nghĩa Phú",
        "Phường Nghĩa Tân",
        "Phường Quảng Thành",
        "Xã Đắk Nia",
        "Xã Đắk R'Moan",
    ],
}


def list_input_workbooks(input_root: Path) -> dict[int, list[Path]]:
    """
    Quét và liệt kê danh sách tệp Excel khảo sát cho các năm 2023 và 2024.
    
    Args:
        input_root (Path): Thư mục gốc chứa các thư mục con 2023, 2024.
        
    Returns:
        dict[int, list[Path]]: Từ điển ánh xạ từ năm đến danh sách đường dẫn tệp.
    """
    result: dict[int, list[Path]] = {}
    for year in [2023, 2024]:
        year_dir = input_root / str(year)
        if not year_dir.exists():
            result[year] = []
            continue
        files = []
        for path in sorted(year_dir.glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            files.append(path)
        result[year] = files
    return result


def read_input_workbook(path: Path) -> pd.DataFrame:
    """
    Đọc tệp Excel đầu vào thành DataFrame của pandas.
    
    Args:
        path (Path): Đường dẫn tệp Excel cần đọc.
        
    Returns:
        pd.DataFrame: DataFrame chứa dữ liệu.
    """
    df = pd.read_excel(path, sheet_name=0, dtype=object, engine="openpyxl")
    return df


def sanitize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Bổ sung các cột bắt buộc bị thiếu trong DataFrame gốc với giá trị None.
    
    Args:
        df (pd.DataFrame): DataFrame gốc.
        
    Returns:
        pd.DataFrame: DataFrame đã được bổ sung cột.
    """
    df = df.copy()
    for col in BASE_INPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df


def template_workbooks(input_root: Path) -> list[Path]:
    """
    Liệt kê danh sách đường dẫn các tệp biểu mẫu báo cáo Excel trong thư mục Format_Report.
    
    Args:
        input_root (Path): Thư mục gốc chứa Format_Report.
        
    Returns:
        list[Path]: Danh sách tệp biểu mẫu Excel.
    """
    report_dir = input_root / "Format_Report"
    if not report_dir.exists():
        return []
    return sorted([p for p in report_dir.glob("*.xlsx") if not p.name.startswith("~$")])


def extract_used_range(ws) -> str | None:
    """
    Tính toán dải ô (used range) có chứa dữ liệu thực sự trong Worksheet Excel.
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        
    Returns:
        str | None: Địa chỉ dải ô (ví dụ: 'A1:H15') hoặc None nếu trống.
    """
    rows = [r for r in ws.iter_rows() if any(c.value is not None and str(c.value).strip() != "" for c in r)]
    if not rows:
        return None
    min_row = min(cell.row for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    max_row = max(cell.row for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    min_col = min(cell.column for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    max_col = max(cell.column for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def extract_report_title(ws) -> str | None:
    """
    Trích xuất tiêu đề báo cáo từ phần đầu của worksheet Excel.
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        
    Returns:
        str | None: Tiêu đề báo cáo đã được làm sạch hoặc None.
    """
    title_keywords = ("TỔNG HỢP", "PHÂN TÍCH", "DANH SÁCH", "TONG HOP", "PHAN TICH", "DANH SACH")
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        values = [v for v in values if v]
        if not values:
            continue
        row_text = " ".join(values)
        upper_text = strip_accents(row_text).upper()
        if any(key in upper_text for key in title_keywords):
            if "UBND" in upper_text or "CỘNG HOÀ" in upper_text or "CỘNG HÒA" in upper_text:
                continue
            return re.sub(r"\s+", " ", row_text).strip()
    return None


def find_title_row(ws) -> tuple[int | None, str | None]:
    """
    Tìm chỉ số dòng chứa tiêu đề báo cáo và nội dung tiêu đề.
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        
    Returns:
        tuple[int | None, str | None]: (chỉ số dòng 1-indexed, nội dung tiêu đề) hoặc (None, None).
    """
    title_keywords = ("TỔNG HỢP", "PHÂN TÍCH", "DANH SÁCH", "TONG HOP", "PHAN TICH", "DANH SACH")
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        values = [v for v in values if v]
        if not values:
            continue
        row_text = " ".join(values)
        upper_text = strip_accents(row_text).upper()
        if any(key in upper_text for key in title_keywords):
            if "UBND" in upper_text or "CONG HOA" in upper_text or "CỘNG HÒA" in upper_text or "CỘNG HOÀ" in upper_text:
                continue
            return row_idx, re.sub(r"\s+", " ", row_text).strip()
    return None, None


def find_first_data_row(ws, start_row: int) -> int | None:
    """
    Tìm dòng đầu tiên bắt đầu chứa dữ liệu cấp hành chính (Huyện/Xã) bên dưới dòng tiêu đề.
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        start_row (int): Chỉ số dòng bắt đầu quét.
        
    Returns:
        int | None: Chỉ số dòng chứa dữ liệu đầu tiên hoặc None.
    """
    for row_idx in range(start_row + 1, min(ws.max_row, start_row + 30) + 1):
        a_val = clean_text(ws.cell(row_idx, 1).value)
        b_val = clean_text(ws.cell(row_idx, 2).value)
        if not a_val or not b_val:
            continue
        a_text = a_val or ""
        b_norm = normalize_key(b_val or "")
        if re.fullmatch(r"\d+", a_text) and (
            b_norm.startswith("huyen") or b_norm.startswith("thanhpho") or b_norm.startswith("xa") or b_norm.startswith("phuong") or b_norm.startswith("thitran")
        ):
            return row_idx
    return None


def extract_header_rows(ws) -> list[int] | None:
    """
    Trích xuất các chỉ số dòng cấu thành tiêu đề bảng dữ liệu (Header rows).
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        
    Returns:
        list[int] | None: Danh sách dòng header hoặc None.
    """
    title_row, _ = find_title_row(ws)
    if title_row is None:
        return None
    first_data_row = find_first_data_row(ws, title_row)
    if first_data_row is None:
        first_data_row = min(ws.max_row, title_row + 8)
    header_rows: list[int] = []
    for row_idx in range(title_row + 1, first_data_row):
        row_values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        row_values = [v for v in row_values if v]
        if not row_values:
            continue
        row_text = " ".join(row_values)
        upper_text = strip_accents(row_text).upper()
        if "UBND" in upper_text or "CONG HOA" in upper_text or "CỘNG HÒA" in upper_text or "CỘNG HOÀ" in upper_text:
            continue
        header_rows.append(row_idx)
    return header_rows or None


def extract_district_commune_mapping(ws) -> dict[str, list[str]]:
    """
    Trích xuất danh sách địa bàn hành chính (các xã tương ứng với từng huyện) từ sheet biểu mẫu.
    
    Args:
        ws: Đối tượng worksheet của openpyxl.
        
    Returns:
        dict[str, list[str]]: Bản đồ Huyện -> [Danh sách Xã].
    """
    mapping: dict[str, list[str]] = {}
    current_district: str | None = None
    for row_idx in range(1, ws.max_row + 1):
        a_val = clean_text(ws.cell(row_idx, 1).value)
        b_val = clean_text(ws.cell(row_idx, 2).value)
        if not a_val and not b_val:
            continue
        a_text = a_val or ""
        b_text = b_val or ""
        b_norm = normalize_key(b_text)
        is_district_row = bool(re.fullmatch(r"\d+", a_text)) and (b_norm.startswith("huyen") or b_norm.startswith("thanhpho"))
        is_commune_row = bool(re.fullmatch(r"\d+\.\d+", a_text)) and (
            b_norm.startswith("xa") or b_norm.startswith("phuong") or b_norm.startswith("thitran")
        )
        if is_district_row:
            current_district = b_text
            mapping.setdefault(current_district, [])
            continue
        if current_district and is_commune_row:
            commune = b_text
            if commune not in mapping[current_district]:
                mapping[current_district].append(commune)
    return mapping


def parse_template_file(path: Path) -> dict[str, Any]:
    """
    Phân tích chi tiết một tệp biểu mẫu báo cáo Excel.
    
    Args:
        path (Path): Đường dẫn tệp Excel biểu mẫu.
        
    Returns:
        dict[str, Any]: Từ điển chứa thông tin cấu trúc biểu mẫu.
    """
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb[wb.sheetnames[0]]
    _, report_name = find_title_row(ws)
    report_name = report_name or path.stem
    header_rows = extract_header_rows(ws)
    district_commune_mapping = extract_district_commune_mapping(ws)
    return {
        "file_name": path.name,
        "report_name": report_name,
        "sheet_name": ws.title,
        "used_range": extract_used_range(ws),
        "header_rows": header_rows[:5] if header_rows else None,
        "district_commune_mapping": district_commune_mapping,
    }


def scan_report_templates(input_root: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]], list[dict[str, Any]]]:
    """
    Quét toàn bộ thư mục Format_Report để phân tích cấu trúc 15 biểu mẫu báo cáo Excel.
    Nếu không có biểu mẫu, sẽ trả về kết quả dự phòng mặc định (fallback).
    
    Args:
        input_root (Path): Thư mục chứa Format_Report.
        
    Returns:
        tuple[list[dict[str, Any]], dict[str, list[str]], list[dict[str, Any]]]: 
            (danh sách schema biểu mẫu, bản đồ Huyện-Xã thu thập được, danh sách cảnh báo).
    """
    template_paths = template_workbooks(input_root)
    warnings: list[dict[str, Any]] = []
    if not template_paths:
        warnings.append(
            {
                "warning_type": "missing_report_templates",
                "year": None,
                "district": None,
                "message": "Format_Report rỗng hoặc không khả dụng; sử dụng cấu hình mặc định (fallback) cho schema và địa bàn hành chính.",
                "severity": "high",
            }
        )
        return build_report_schema_summary(), DISTRICT_COMMUNE_FALLBACK.copy(), warnings

    parsed = [parse_template_file(path) for path in template_paths]
    mapping: dict[str, list[str]] = {}
    for item in parsed:
        for district, communes in item["district_commune_mapping"].items():
            mapping.setdefault(district, [])
            for commune in communes:
                if commune not in mapping[district]:
                    mapping[district].append(commune)
                    
    # Nếu bản đồ quét từ template trống hoàn toàn, dùng fallback
    if not mapping or not any(mapping.values()):
        mapping = DISTRICT_COMMUNE_FALLBACK.copy()

    summary = build_report_schema_summary(template_status="parsed_from_templates")
    for item in summary:
        for parsed_item in parsed:
            if parsed_item["file_name"].startswith(f"{item['report_id']}."):
                item["sheet_name"] = parsed_item["sheet_name"]
                item["used_range"] = parsed_item["used_range"]
                item["header_rows"] = parsed_item["header_rows"]
                item["report_name"] = parsed_item["report_name"]
                break
    return summary, mapping, warnings
