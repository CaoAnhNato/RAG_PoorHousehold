from __future__ import annotations

import datetime as dt
import hashlib
import json
import math
import random
import re
import unicodedata
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


CORE_COLUMNS = [
    "date",
    "family.hostName",
    "family.code",
    "family.coDanTocTaiCho",
    "family.hostGender",
    "family.numberOfMembers",
    "quickReview.result",
    "b1Point",
    "b2Point",
    "classify",
]

BASE_INPUT_COLUMNS = [
    "date",
    "family.hostName",
    "family.code",
    "family.coDanTocTaiCho",
    "family.hostGender",
    "family.numberOfMembers",
    "quickReview.result",
    "b1Point",
    "b2Point",
    "classify",
    "reviewer",
]

DATE_OPTIONS_2023 = ["29/10/2023", "05/11/2023", "12/11/2023", "19/11/2023"]
DATE_OPTIONS_2024 = ["29/10/2024", "05/11/2024", "12/11/2024", "19/11/2024"]

ETHNICITY_POOL = [
    "Kinh",
    "Ê đê",
    "Mạ",
    "Mường",
    "Thái",
    "M'Nông",
    "Tày",
    "Nùng",
    "Mông",
    "Dao",
    "DT Khác",
]

DISTRICT_COMMUNE_FALLBACK = {
    "Huyện Cư Jút": [
        "Thị trấn Ea T'ling",
        "Xã Cư Knia",
        "Xã Đắk DRông",
        "Xã Đắk Wil",
        "Xã Ea Pô",
        "Xã Nam Dong",
        "Xã Tâm Thắng",
        "Xã Trúc Sơn",
    ],
    "Huyện Krông Nô": [
        "Thị trấn Đắk Mâm",
        "Xã Buôn Choáh",
        "Xã Đắk Drô",
        "Xã Đắk N'Drót",
        "Xã Đắk Sôr",
        "Xã Đức Xuyên",
        "Xã Nam Đà",
        "Xã Nam Xuân",
        "Xã Nâm N'Đir",
        "Xã Nâm Nung",
        "Xã Quảng Phú",
        "Xã Tân Thành",
    ],
    "Huyện Tuy Đức": [
        "Xã Đắk Búk So",
        "Xã Đắk Ngo",
        "Xã Đắk R'Tih",
        "Xã Quảng Tâm",
        "Xã Quảng Tân",
        "Xã Quảng Trực",
    ],
    "Huyện Đăk Glong": [
        "Xã Quảng Khê",
        "Xã Quảng Sơn",
        "Xã Quảng Hòa",
        "Xã Đắk Ha",
        "Xã Đắk Som",
        "Xã Đắk R'Măng",
        "Xã Đắk Plao",
    ],
    "Huyện Đắk Mil": [
        "Thị trấn Đắk Mil",
        "Xã Đắk Gằn",
        "Xã Đắk Lao",
        "Xã Đắk N'Drót",
        "Xã Đắk R'La",
        "Xã Đắk Sắk",
        "Xã Đức Mạnh",
        "Xã Đức Minh",
        "Xã Long Sơn",
        "Xã Thuận An",
    ],
    "Huyện Đắk RLấp": [
        "Thị trấn Kiến Đức",
        "Xã Nhân Cơ",
        "Xã Đắk Wer",
        "Xã Kiến Thành",
        "Xã Đạo Nghĩa",
        "Xã Nghĩa Thắng",
        "Xã Nhân Đạo",
        "Xã Đắk Sin",
        "Xã Quảng Tín",
        "Xã Hưng Bình",
        "Xã Đắk Ru",
    ],
    "Huyện Đắk Song": [
        "Thị trấn Đức An",
        "Xã Đắk Hòa",
        "Xã Đắk Môl",
        "Xã Đắk N'Drung",
        "Xã Nam Bình",
        "Xã Nâm N'Jang",
        "Xã Thuận Hà",
        "Xã Thuận Hạnh",
        "Xã Trường Xuân",
    ],
    "Thành phố Gia Nghĩa": [
        "Phường Nghĩa Đức",
        "Phường Nghĩa Thành",
        "Phường Nghĩa Trung",
        "Phường Nghĩa Phú",
        "Phường Nghĩa Tân",
        "Phường Quảng Thành",
        "Xã Đắk Nia",
        "Xã Đắk R'Moan",
    ],
}

DISTRICT_MINORITY_PROFILE = {
    "Huyện Cư Jút": "high",
    "Huyện Krông Nô": "high",
    "Huyện Tuy Đức": "high",
    "Huyện Đăk Glong": "high",
    "Huyện Đắk Mil": "medium",
    "Huyện Đắk RLấp": "medium",
    "Huyện Đắk Song": "medium",
    "Thành phố Gia Nghĩa": "urban",
}

LEGAL_PERIOD = "2021_2025"

DATE_OPTIONS_BY_YEAR = {
    2023: ["2023-09-15", "2023-10-15", "2023-11-15", "2023-12-01"],
    2024: ["2024-09-15", "2024-10-15", "2024-11-15", "2024-12-01"],
}

LOCAL_ETHNICITY_MAPPING = {
    "province": "Đắk Nông",
    "source_note": "Cấu hình dự án cần được địa phương kiểm tra lại trước khi dùng cho phân tích chính sách chính thức.",
    "default_local_ethnicities": ["M'Nông", "Mạ", "Ê Đê"],
    "district_overrides": {
        "Huyện Cư Jút": ["M'Nông", "Ê Đê"],
        "Huyện Đắk Mil": ["M'Nông", "Ê Đê", "Mạ"],
        "Huyện Đắk Song": ["M'Nông", "Mạ"],
        "Huyện Đắk RLấp": ["M'Nông", "Mạ"],
        "Huyện Đăk Glong": ["M'Nông", "Mạ"],
        "Huyện Tuy Đức": ["M'Nông", "Mạ"],
        "Huyện Krông Nô": ["M'Nông", "Ê Đê"],
        "Thành phố Gia Nghĩa": ["M'Nông", "Mạ", "Ê Đê"],
    },
    "confidence": "controlled_project_assumption_needs_local_review",
}

VILLAGE_CANDIDATES = [
    "Thôn 1",
    "Thôn 2",
    "Thôn 3",
    "Thôn 4",
    "Bon 1",
    "Bon 2",
    "Bon 3",
    "Tổ dân phố 1",
    "Tổ dân phố 2",
    "Tổ dân phố 3",
]

NON_LOCAL_ETHNICITIES = ["Kinh", "Tày", "Nùng", "Mường", "Thái", "Mông", "Dao", "DT Khác"]


def is_missingish(value: Any) -> bool:
    text = clean_text(value)
    return text is None or text in {"", "None", "NaN", "nan"}


def audit_suffix(column_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", column_name).strip("_")


def candidate_dates_for_year(year: int) -> list[str]:
    return DATE_OPTIONS_BY_YEAR.get(year, DATE_OPTIONS_BY_YEAR[2024])


def resolve_area_type_info(commune: Any) -> tuple[str, str, str]:
    commune_text = clean_text(commune) or ""
    normalized = strip_accents(commune_text).lower().strip()
    if normalized.startswith("xa"):
        return "rural", "generated_by_legal_rule", "high"
    if normalized.startswith("phuong") or normalized.startswith("thi tran"):
        return "urban", "generated_by_legal_rule", "high"
    return "rural", "template_required_default", "low"


def review_period_for_year(year: int) -> str:
    return f"{year}_annual_review"


def resolve_review_type(date_value: Any) -> str:
    parsed = parse_date_value(date_value)
    if parsed is None:
        return "Rà soát định kỳ hằng năm"
    if dt.date(parsed.year, 9, 1) <= parsed <= dt.date(parsed.year, 12, 14):
        return "Rà soát định kỳ hằng năm"
    if dt.date(parsed.year, 1, 1) <= parsed <= dt.date(parsed.year, 8, 31):
        return "Rà soát trong năm"
    return "Rà soát đột xuất/cuối năm"


def resolve_quick_review_result(value: Any) -> tuple[str, str, bool]:
    text = clean_text(value)
    if text is None or text == "--":
        return "Đủ điều kiện rà soát B1/B2", "template_required_default", True
    return re.sub(r"\s+", " ", text).strip(), "cleaned_original", False


def resolve_reviewer_value(value: Any) -> tuple[str, str, bool]:
    text = clean_text(value)
    if text is None:
        return "Rà soát viên cấp xã", "template_required_default", True
    return re.sub(r"\s+", " ", text).strip(), "cleaned_original", False


def resolve_host_name_value(value: Any, gender: str, ethnicity: str, rng: random.Random) -> tuple[str, str, bool]:
    text = clean_text(value)
    if text:
        return re.sub(r"\s+", " ", text).strip(), "cleaned_original", False
    name = generate_member_name(rng, gender, ethnicity)
    return name, "generated_by_controlled_synthetic_rule", True


def resolve_family_members_count(value: Any, classify: str, rng: random.Random) -> tuple[int, str, bool]:
    size = parse_int(value, default=0)
    if size > 0:
        return size, "cleaned_original", False
    if classify == "Hộ nghèo":
        return rng.randint(2, 7), "generated_by_controlled_synthetic_rule", True
    if classify == "Hộ cận nghèo":
        return rng.randint(2, 6), "generated_by_controlled_synthetic_rule", True
    return rng.randint(1, 6), "generated_by_controlled_synthetic_rule", True


def resolve_ethnicity_and_dttc(
    district: str,
    classify: str,
    original_co_dan_toc: Any,
    rng: random.Random,
) -> tuple[str, str, bool]:
    text = clean_text(original_co_dan_toc)
    local_ethnicities = LOCAL_ETHNICITY_MAPPING["district_overrides"].get(district, LOCAL_ETHNICITY_MAPPING["default_local_ethnicities"])
    if text is None:
        if classify == "Hộ nghèo":
            co_dan_toc = "Có" if rng.random() < 0.72 else "Không"
        elif classify == "Hộ cận nghèo":
            co_dan_toc = "Có" if rng.random() < 0.48 else "Không"
        else:
            co_dan_toc = "Có" if rng.random() < 0.18 else "Không"
    else:
        co_dan_toc = normalize_yes_no(text)

    if co_dan_toc == "Có":
        ethnicity = rng.choice(local_ethnicities)
        if ethnicity == "Kinh":
            ethnicity = rng.choice([e for e in local_ethnicities if e != "Kinh"] or local_ethnicities)
        return ethnicity, "Có", True

    ethnicity = rng.choice([e for e in NON_LOCAL_ETHNICITIES if e != "Kinh"])
    return ethnicity, "Không", False


def resolve_b1_point(classify: str, area_type: str, original_value: Any, rng: random.Random) -> tuple[int, str, bool]:
    original_int = parse_int(original_value, default=0)
    if area_type == "urban":
        ranges = {
            "Hộ nghèo": (90, 175),
            "Hộ cận nghèo": (110, 175),
            "Hộ không nghèo": (176, 260),
        }
    else:
        ranges = {
            "Hộ nghèo": (60, 140),
            "Hộ cận nghèo": (80, 140),
            "Hộ không nghèo": (141, 210),
        }
    low, high = ranges.get(classify, ranges["Hộ không nghèo"])
    generated = rng.randint(low, high)
    if original_int > 0 and low <= original_int <= high:
        return original_int, "cleaned_original", False
    return generated, "generated_by_legal_rule", True


def resolve_deprivation_total_count(
    classify: str,
    original_b2_value: Any,
    original_total_value: Any,
    rng: random.Random,
) -> tuple[int, int, str, bool]:
    original_b2 = parse_int(original_b2_value, default=-1)
    original_total = parse_int(original_total_value, default=-1)
    if classify == "Hộ nghèo":
        total = rng.randint(3, 8)
    else:
        total = rng.randint(0, 2)

    if original_total >= 0 and 0 <= original_total <= 12:
        total = original_total
    elif original_b2 in list(range(0, 121, 10)):
        total = int(original_b2 / 10)
    elif 0 <= original_b2 <= 12:
        total = original_b2

    if classify == "Hộ nghèo":
        total = max(3, min(8, total))
    else:
        total = max(0, min(2, total))

    return total, total * 10, "generated_by_legal_rule", True


def resolve_classification_from_scores(b1_point: int, deprivation_total: int, area_type: str) -> str:
    threshold = 175 if area_type == "urban" else 140
    if b1_point <= threshold and deprivation_total >= 3:
        return "Hộ nghèo"
    if b1_point <= threshold and deprivation_total < 3:
        return "Hộ cận nghèo"
    return "Hộ không nghèo"


def resolve_transition_change_types(beginning: str, ending: str) -> tuple[str, str, bool]:
    if beginning == "Hộ nghèo" and ending == "Hộ nghèo":
        return "poor_beginning", "none", False
    if beginning == "Hộ nghèo" and ending == "Hộ cận nghèo":
        return "poor_to_near_poor", "poor_to_near_poor", False
    if beginning == "Hộ nghèo" and ending == "Hộ không nghèo":
        return "poor_escape_above_near_poor", "none", True
    if beginning == "Hộ cận nghèo" and ending == "Hộ nghèo":
        return "near_poor_to_poor", "near_poor_to_poor", False
    if beginning == "Hộ cận nghèo" and ending == "Hộ cận nghèo":
        return "none", "near_poor_beginning", False
    if beginning == "Hộ cận nghèo" and ending == "Hộ không nghèo":
        return "none", "near_poor_escape", True
    if beginning == "Hộ không nghèo" and ending == "Hộ nghèo":
        return "new_poor", "none", False
    if beginning == "Hộ không nghèo" and ending == "Hộ cận nghèo":
        return "none", "new_near_poor", False
    return "none", "none", False


REPORT_SPECS = [
    {
        "report_id": 1,
        "report_name": "Tổng hợp kết quả rà soát HC,CN,NL,NN có mức sống trung bình",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Tổng hợp theo huyện", "Địa bàn xã/phường/thị trấn"],
        "required_columns": [
            "administrative.province",
            "administrative.district",
            "administrative.commune",
            "family.isAgricultureForestryFisherySaltMediumLivingStandard",
            "classify",
        ],
        "notes": {
            "family.isAgricultureForestryFisherySaltMediumLivingStandard": "aggregate count / flag",
            "classify": "grouping label",
        },
    },
    {
        "report_id": 2,
        "report_name": "Tổng hợp diễn biến hộ nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Biến động hộ nghèo", "Đầu kỳ / cuối kỳ"],
        "required_columns": [
            "transition.beginningClassify",
            "transition.endingClassify",
            "transition.poorChangeType",
            "classify",
            "administrative.district",
            "administrative.commune",
        ],
        "notes": {
            "transition.beginningClassify": "detail list / aggregate count",
            "transition.endingClassify": "detail list / aggregate count",
            "transition.poorChangeType": "aggregate count",
        },
    },
    {
        "report_id": 3,
        "report_name": "Tổng hợp diễn biến hộ cận nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Biến động hộ cận nghèo", "Đầu kỳ / cuối kỳ"],
        "required_columns": [
            "transition.beginningClassify",
            "transition.endingClassify",
            "transition.nearPoorChangeType",
            "classify",
            "administrative.district",
            "administrative.commune",
        ],
        "notes": {
            "transition.beginningClassify": "detail list / aggregate count",
            "transition.endingClassify": "detail list / aggregate count",
            "transition.nearPoorChangeType": "aggregate count",
        },
    },
    {
        "report_id": 4,
        "report_name": "Phân tích chỉ số thiếu hụt dịch vụ xã hội cơ bản hộ nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["12 chỉ số thiếu hụt", "Tổng số thiếu hụt"],
        "required_columns": [
            "classify",
            "deprivation.totalCount",
            "deprivation.employment",
            "deprivation.dependentPerson",
            "deprivation.nutrition",
            "deprivation.healthInsurance",
            "deprivation.adultEducation",
            "deprivation.childSchoolAttendance",
            "deprivation.housingQuality",
            "deprivation.housingArea",
            "deprivation.cleanWater",
            "deprivation.hygienicToilet",
            "deprivation.telecommunication",
            "deprivation.informationAccessAssets",
        ],
        "notes": {f"deprivation.{name}": "aggregate count" for name in [
            "employment",
            "dependentPerson",
            "nutrition",
            "healthInsurance",
            "adultEducation",
            "childSchoolAttendance",
            "housingQuality",
            "housingArea",
            "cleanWater",
            "hygienicToilet",
            "telecommunication",
            "informationAccessAssets",
        ]} | {"deprivation.totalCount": "sum / aggregate count"},
    },
    {
        "report_id": 5,
        "report_name": "Phân tích tỷ lệ các chỉ số thiếu hụt dịch vụ xã hội cơ bản hộ nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["12 chỉ số thiếu hụt", "Tỷ lệ phần trăm"],
        "required_columns": [
            "classify",
            "deprivation.totalCount",
            "deprivation.employment",
            "deprivation.dependentPerson",
            "deprivation.nutrition",
            "deprivation.healthInsurance",
            "deprivation.adultEducation",
            "deprivation.childSchoolAttendance",
            "deprivation.housingQuality",
            "deprivation.housingArea",
            "deprivation.cleanWater",
            "deprivation.hygienicToilet",
            "deprivation.telecommunication",
            "deprivation.informationAccessAssets",
        ],
        "notes": {f"deprivation.{name}": "percentage" for name in [
            "employment",
            "dependentPerson",
            "nutrition",
            "healthInsurance",
            "adultEducation",
            "childSchoolAttendance",
            "housingQuality",
            "housingArea",
            "cleanWater",
            "hygienicToilet",
            "telecommunication",
            "informationAccessAssets",
        ]},
    },
    {
        "report_id": 6,
        "report_name": "Phân tích chỉ số thiếu hụt dịch vụ xã hội cơ bản hộ cận nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["12 chỉ số thiếu hụt", "Tổng số thiếu hụt"],
        "required_columns": [
            "classify",
            "deprivation.totalCount",
            "deprivation.employment",
            "deprivation.dependentPerson",
            "deprivation.nutrition",
            "deprivation.healthInsurance",
            "deprivation.adultEducation",
            "deprivation.childSchoolAttendance",
            "deprivation.housingQuality",
            "deprivation.housingArea",
            "deprivation.cleanWater",
            "deprivation.hygienicToilet",
            "deprivation.telecommunication",
            "deprivation.informationAccessAssets",
        ],
        "notes": {f"deprivation.{name}": "aggregate count" for name in [
            "employment",
            "dependentPerson",
            "nutrition",
            "healthInsurance",
            "adultEducation",
            "childSchoolAttendance",
            "housingQuality",
            "housingArea",
            "cleanWater",
            "hygienicToilet",
            "telecommunication",
            "informationAccessAssets",
        ]} | {"deprivation.totalCount": "sum / aggregate count"},
    },
    {
        "report_id": 7,
        "report_name": "Phân tích tỷ lệ các chỉ số thiếu hụt dịch vụ xã hội cơ bản hộ cận nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["12 chỉ số thiếu hụt", "Tỷ lệ phần trăm"],
        "required_columns": [
            "classify",
            "deprivation.totalCount",
            "deprivation.employment",
            "deprivation.dependentPerson",
            "deprivation.nutrition",
            "deprivation.healthInsurance",
            "deprivation.adultEducation",
            "deprivation.childSchoolAttendance",
            "deprivation.housingQuality",
            "deprivation.housingArea",
            "deprivation.cleanWater",
            "deprivation.hygienicToilet",
            "deprivation.telecommunication",
            "deprivation.informationAccessAssets",
        ],
        "notes": {f"deprivation.{name}": "percentage" for name in [
            "employment",
            "dependentPerson",
            "nutrition",
            "healthInsurance",
            "adultEducation",
            "childSchoolAttendance",
            "housingQuality",
            "housingArea",
            "cleanWater",
            "hygienicToilet",
            "telecommunication",
            "informationAccessAssets",
        ]},
    },
    {
        "report_id": 8,
        "report_name": "Phân tích hộ nghèo, hộ cận nghèo theo các nhóm đối tượng",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Nhóm đối tượng", "Hộ nghèo / cận nghèo"],
        "required_columns": [
            "classify",
            "family.hasNoLaborCapacity",
            "family.hasRevolutionMeritPolicy",
            "family.isDTTC",
            "support.health",
            "support.education",
            "support.production",
            "support.credit",
            "support.housing",
            "support.other",
        ],
        "notes": {
            "family.hasNoLaborCapacity": "aggregate count",
            "family.hasRevolutionMeritPolicy": "aggregate count",
            "family.isDTTC": "aggregate count",
            "support.health": "aggregate count",
            "support.education": "aggregate count",
            "support.production": "aggregate count",
            "support.credit": "aggregate count",
            "support.housing": "aggregate count",
            "support.other": "aggregate count",
        },
    },
    {
        "report_id": 9,
        "report_name": "Phân tích hộ nghèo, hộ cận nghèo theo dân tộc",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Dân tộc", "Hộ nghèo / cận nghèo"],
        "required_columns": [
            "family.ethnicity",
            "family.isDTTS",
            "family.isKinh",
            "classify",
        ],
        "notes": {
            "family.ethnicity": "detail list / aggregate count",
            "family.isDTTS": "aggregate count",
            "family.isKinh": "aggregate count",
        },
    },
    {
        "report_id": 10,
        "report_name": "Phân tích hộ nghèo, hộ cận nghèo theo nguyên nhân nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Nguyên nhân nghèo/cận nghèo", "Hộ nghèo / cận nghèo"],
        "required_columns": [
            "reason.lackProductionLand",
            "reason.lackCapital",
            "reason.lackLabor",
            "reason.lackProductionTools",
            "reason.lackProductionKnowledge",
            "reason.lackLaborSkill",
            "reason.illnessOrAccident",
            "reason.other",
            "classify",
        ],
        "notes": {f"reason.{name}": "aggregate count" for name in [
            "lackProductionLand",
            "lackCapital",
            "lackLabor",
            "lackProductionTools",
            "lackProductionKnowledge",
            "lackLaborSkill",
            "illnessOrAccident",
            "other",
        ]},
    },
    {
        "report_id": 11,
        "report_name": "Tổng hợp chỉ tiêu thiếu hụt của trẻ em thuộc hộ nghèo, hộ cận nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Trẻ em", "Thiếu hụt của trẻ em"],
        "required_columns": [
            "children.totalCount",
            "children.lackHealthInsuranceCount",
            "children.nutritionDeprivedCount",
            "children.schoolAttendanceDeprivedCount",
            "classify",
        ],
        "notes": {
            "children.totalCount": "sum",
            "children.lackHealthInsuranceCount": "sum",
            "children.nutritionDeprivedCount": "sum",
            "children.schoolAttendanceDeprivedCount": "sum",
        },
    },
    {
        "report_id": 12,
        "report_name": "Tổng hợp kết quả rà soát hộ nghèo theo chuẩn đa chiều",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Chuẩn đa chiều", "Hộ nghèo"],
        "required_columns": [
            "classify",
            "family.povertyStatusDetail",
            "family.hasNoLaborCapacity",
            "family.hasRevolutionMeritPolicy",
            "family.ethnicity",
            "family.isDTTC",
            "support.health",
            "support.education",
            "support.production",
            "support.credit",
            "support.housing",
            "support.other",
            "deprivation.totalCount",
            "children.totalCount",
        ],
        "notes": {
            "family.povertyStatusDetail": "detail list / aggregate count",
            "deprivation.totalCount": "sum",
            "children.totalCount": "sum",
        },
    },
    {
        "report_id": 13,
        "report_name": "Tổng hợp kết quả rà soát hộ cận nghèo theo chuẩn đa chiều",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Chuẩn đa chiều", "Hộ cận nghèo"],
        "required_columns": [
            "classify",
            "family.nearPovertyStatusDetail",
            "family.hasNoLaborCapacity",
            "family.hasRevolutionMeritPolicy",
            "family.ethnicity",
            "family.isDTTC",
            "support.health",
            "support.education",
            "support.production",
            "support.credit",
            "support.housing",
            "support.other",
            "deprivation.totalCount",
            "children.totalCount",
        ],
        "notes": {
            "family.nearPovertyStatusDetail": "detail list / aggregate count",
            "deprivation.totalCount": "sum",
            "children.totalCount": "sum",
        },
    },
    {
        "report_id": 14,
        "report_name": "Danh sách chi tiết hộ cận nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Danh sách hộ cận nghèo", "Danh sách thành viên hộ"],
        "required_columns": [
            "family.membersGenerated",
            "family.membersFile",
            "family.membersJson",
            "family.nearPovertyStatusDetail",
            "family.hasRevolutionMeritPolicy",
            "family.isDTTC",
            "family.hasNoLaborCapacity",
            "support.health",
            "support.education",
            "support.production",
            "support.credit",
            "support.housing",
            "support.other",
        ],
        "notes": {
            "family.membersGenerated": "detail list",
            "family.membersFile": "detail list",
            "family.membersJson": "detail list",
        },
    },
    {
        "report_id": 15,
        "report_name": "Danh sách chi tiết hộ nghèo",
        "sheet_name": None,
        "used_range": None,
        "header_rows": None,
        "main_groups": ["Danh sách hộ nghèo", "Danh sách thành viên hộ"],
        "required_columns": [
            "family.membersGenerated",
            "family.membersFile",
            "family.membersJson",
            "family.povertyStatusDetail",
            "family.hasRevolutionMeritPolicy",
            "family.isDTTC",
            "family.hasNoLaborCapacity",
            "support.health",
            "support.education",
            "support.production",
            "support.credit",
            "support.housing",
            "support.other",
        ],
        "notes": {
            "family.membersGenerated": "detail list",
            "family.membersFile": "detail list",
            "family.membersJson": "detail list",
        },
    },
]


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def strip_accents(text: str) -> str:
    if text is None:
        return ""
    text = unicodedata.normalize("NFKD", str(text))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_key(text: str) -> str:
    text = strip_accents(text).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def normalize_no_space(text: str) -> str:
    return re.sub(r"\s+", "", strip_accents(str(text)).lower())


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def clean_code(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    text = text.replace("\u00a0", " ")
    text = text.strip()
    if text.endswith(".0"):
        try:
            num = int(float(text))
            return str(num)
        except Exception:
            text = text[:-2]
    if text.endswith(".") and re.fullmatch(r"\d+\.", text):
        return text[:-1]
    if re.fullmatch(r"\d+\.0", text):
        return str(int(float(text)))
    return text


def parse_int(value: Any, default: int = 1) -> int:
    if value is None:
        return default
    if isinstance(value, (int, np.integer)):
        return max(default, int(value))
    if isinstance(value, float) and not math.isnan(value):
        return max(default, int(value))
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return default
    text = text.replace(",", ".")
    try:
        return max(default, int(float(text)))
    except Exception:
        digits = re.findall(r"\d+", text)
        if digits:
            return max(default, int(digits[0]))
        return default


def stable_seed(*parts: Any) -> int:
    joined = "|".join("" if p is None else str(p) for p in parts)
    digest = hashlib.sha256(joined.encode("utf-8")).hexdigest()
    return int(digest[:16], 16)


def rng_for(*parts: Any) -> random.Random:
    return random.Random(stable_seed(*parts))


def normalize_yes_no(value: Any) -> str:
    text = clean_text(value)
    if text is None:
        return "Không"
    lowered = strip_accents(text).lower().strip()
    true_values = {"co", "có", "1", "true", "yes", "y", "t"}
    false_values = {"khong", "không", "0", "false", "no", "n", "f"}
    if lowered in true_values:
        return "Có"
    if lowered in false_values:
        return "Không"
    if "co" == lowered or lowered.startswith("co "):
        return "Có"
    if "khong" == lowered or lowered.startswith("khong "):
        return "Không"
    return "Không"


def normalize_gender(value: Any, name: str | None = None, rng: random.Random | None = None) -> str:
    text = clean_text(value)
    if text is not None:
        lowered = strip_accents(text).lower()
        if lowered.startswith("nam") or lowered in {"m", "male"}:
            return "Nam"
        if lowered.startswith("nu") or lowered.startswith("nữ") or lowered in {"f", "female"}:
            return "Nữ"
    if name:
        female_tokens = ["thi", "thị", "ngoc", "ngọc", "hương", "huong", "anh", "lan", "mai", "ly", "le", "lệ", "nga", "hoa", "hồng", "hong"]
        male_tokens = ["van", "văn", "duc", "đức", "quang", "tuan", "tùng", "tung", "hieu", "hiếu", "minh", "dung", "hoang", "hùng", "hung"]
        normalized_name = strip_accents(name).lower()
        female_score = sum(1 for tok in female_tokens if tok in normalized_name)
        male_score = sum(1 for tok in male_tokens if tok in normalized_name)
        if female_score > male_score:
            return "Nữ"
        if male_score > female_score:
            return "Nam"
    if rng is None:
        rng = random.Random()
    return "Nam" if rng.random() < 0.5 else "Nữ"


def parse_date_value(value: Any) -> dt.date | None:
    text = clean_text(value)
    if text is None:
        return None
    candidates = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in candidates:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def format_date(date_value: dt.date | None) -> str | None:
    if date_value is None:
        return None
    return date_value.strftime("%Y-%m-%d")


def set_year(date_value: dt.date | None, year: int) -> dt.date | None:
    if date_value is None:
        return None
    try:
        return dt.date(year, date_value.month, date_value.day)
    except ValueError:
        return None


def maybe_random_date(year: int, rng: random.Random) -> str:
    if year == 2023:
        return rng.choice(DATE_OPTIONS_2023)
    return rng.choice(DATE_OPTIONS_2024)


def is_header_row(row: pd.Series) -> bool:
    classify = clean_text(row.get("classify"))
    host = clean_text(row.get("family.hostName"))
    date_val = clean_text(row.get("date"))
    return classify == "Kết quả rà soát" or host == "Chủ hộ" or date_val == "Thời gian rà soát"


def classify_bucket(classify: str | None) -> str:
    if classify == "Hộ nghèo":
        return "poor"
    if classify == "Hộ cận nghèo":
        return "near_poor"
    return "other"


def resolve_district_communes(district_name: str, commune_mapping: dict[str, list[str]] | None = None) -> list[str]:
    source = commune_mapping or DISTRICT_COMMUNE_FALLBACK
    if district_name in source:
        return source[district_name]
    normalized = normalize_key(district_name)
    for key, communes in source.items():
        if normalize_key(key) == normalized:
            return communes
    return []


def weighted_choice(rng: random.Random, weighted_items: list[tuple[Any, float]]) -> Any:
    total = sum(max(0.0, w) for _, w in weighted_items)
    if total <= 0:
        return weighted_items[0][0]
    pick = rng.random() * total
    cumulative = 0.0
    for item, weight in weighted_items:
        cumulative += max(0.0, weight)
        if pick <= cumulative:
            return item
    return weighted_items[-1][0]


def choose_from_weights(rng: random.Random, values: list[Any], weights: list[float]) -> Any:
    return weighted_choice(rng, list(zip(values, weights)))


def assign_communes(communes: list[str], n: int, rng: random.Random) -> list[str]:
    if not communes:
        return [None] * n
    if n <= len(communes):
        ordered = communes[:]
        rng.shuffle(ordered)
        return ordered[:n]
    base = communes[:]
    rng.shuffle(base)
    repeated: list[str] = []
    while len(repeated) < n:
        chunk = base[:]
        rng.shuffle(chunk)
        repeated.extend(chunk)
    return repeated[:n]


def village_groups_for_n(n: int) -> list[str]:
    return [
        "Thôn 1",
        "Thôn 2",
        "Thôn 3",
        "Bon 1",
        "Bon 2",
        "Tổ dân phố 1",
        "Tổ dân phố 2",
    ][: max(1, min(n, 7))]


def assign_village_or_group(rng: random.Random) -> str:
    return rng.choice([
        "Thôn 1",
        "Thôn 2",
        "Thôn 3",
        "Bon 1",
        "Bon 2",
        "Tổ dân phố 1",
        "Tổ dân phố 2",
    ])


def district_ethnicity_weights(district: str, is_dttc: bool) -> list[tuple[str, float]]:
    profile = DISTRICT_MINORITY_PROFILE.get(district, "medium")
    if is_dttc:
        return [
            ("M'Nông", 0.35),
            ("Mạ", 0.20),
            ("Ê đê", 0.15),
            ("Mường", 0.06),
            ("Thái", 0.05),
            ("Tày", 0.05),
            ("Nùng", 0.05),
            ("Mông", 0.04),
            ("Dao", 0.03),
            ("DT Khác", 0.02),
            ("Kinh", 0.0),
        ]
    if profile == "high":
        return [
            ("Kinh", 0.35),
            ("M'Nông", 0.20),
            ("Ê đê", 0.10),
            ("Mạ", 0.08),
            ("Mường", 0.08),
            ("Thái", 0.05),
            ("Tày", 0.05),
            ("Nùng", 0.04),
            ("Mông", 0.03),
            ("Dao", 0.03),
            ("DT Khác", 0.09),
        ]
    if profile == "urban":
        return [
            ("Kinh", 0.70),
            ("M'Nông", 0.08),
            ("Ê đê", 0.06),
            ("Mạ", 0.04),
            ("Mường", 0.03),
            ("Thái", 0.03),
            ("Tày", 0.03),
            ("Nùng", 0.01),
            ("Mông", 0.01),
            ("Dao", 0.01),
            ("DT Khác", 0.00),
        ]
    return [
        ("Kinh", 0.55),
        ("M'Nông", 0.12),
        ("Ê đê", 0.08),
        ("Mạ", 0.06),
        ("Mường", 0.05),
        ("Thái", 0.04),
        ("Tày", 0.04),
        ("Nùng", 0.03),
        ("Mông", 0.02),
        ("Dao", 0.01),
        ("DT Khác", 0.00),
    ]


def choose_ethnicity(district: str, is_dttc: bool, rng: random.Random) -> str:
    choices = district_ethnicity_weights(district, is_dttc)
    values = [item for item, _ in choices]
    weights = [weight for _, weight in choices]
    return choose_from_weights(rng, values, weights)


def age_to_birthdate(age: int, rng: random.Random, year_anchor: int) -> dt.date:
    age = max(0, age)
    birth_year = year_anchor - age
    month = rng.randint(1, 12)
    day = rng.randint(1, 28)
    return dt.date(birth_year, month, day)


def host_age_for_row(row: pd.Series, rng: random.Random) -> int:
    classify = row.get("classify")
    size = parse_int(row.get("family.numberOfMembers"), default=1)
    if classify == "Hộ nghèo":
        return rng.randint(30, 78 if size > 1 else 85)
    if classify == "Hộ cận nghèo":
        return rng.randint(28, 80)
    if classify == "Hộ không nghèo":
        return rng.randint(25, 75)
    return rng.randint(25, 80)


def allocate_member_ages(size: int, classify: str | None, host_age: int, rng: random.Random) -> list[int]:
    if size <= 1:
        return [host_age]
    bucket = classify_bucket(classify)
    child_prob = {"poor": 0.42, "near_poor": 0.34, "other": 0.22}[bucket]
    elder_prob = {"poor": 0.10, "near_poor": 0.08, "other": 0.12}[bucket]
    ages = [host_age]
    for _ in range(size - 1):
        roll = rng.random()
        if roll < child_prob:
            ages.append(rng.randint(0, 15))
        elif roll < child_prob + elder_prob:
            ages.append(rng.randint(60, 90))
        else:
            low = max(16, min(45, host_age - 20))
            high = max(low + 5, min(59, host_age + 15))
            ages.append(rng.randint(low, high))
    return ages


def member_name_pool(gender: str, ethnicity: str) -> tuple[list[str], list[str]]:
    male = [
        "Nguyễn Văn An",
        "Trần Văn Bình",
        "Lê Văn Cường",
        "Phạm Văn Đức",
        "Hoàng Văn Dũng",
        "Phan Văn Hải",
        "Võ Văn Hùng",
        "Đặng Văn Khoa",
        "Bùi Văn Lâm",
        "Đỗ Văn Minh",
        "Ngô Văn Nam",
        "Dương Văn Phúc",
    ]
    female = [
        "Nguyễn Thị Mai",
        "Trần Thị Lan",
        "Lê Thị Hạnh",
        "Phạm Thị Hoa",
        "Hoàng Thị Ngọc",
        "Phan Thị Hương",
        "Võ Thị Loan",
        "Đặng Thị Nhung",
        "Bùi Thị Phượng",
        "Đỗ Thị Thanh",
        "Ngô Thị Thúy",
        "Dương Thị Vân",
    ]
    if ethnicity != "Kinh":
        male = [
            "Y Phúc",
            "Y Lộc",
            "Y Sang",
            "H'Rin",
            "H'My",
            "A Mí",
            "A Kha",
            "Rơ Mah Đức",
            "K'Tiêng",
            "Siu Khoa",
        ] + male
        female = [
            "H'Lan",
            "H'Nhung",
            "Y Lanh",
            "Y Nhi",
            "H'Mây",
            "A Dung",
            "Rơ Chăm Mai",
            "K'Ly",
            "Siu Hồng",
            "A Nhiên",
        ] + female
    return male, female


def generate_member_name(rng: random.Random, gender: str, ethnicity: str) -> str:
    male_pool, female_pool = member_name_pool(gender, ethnicity)
    if gender == "Nữ":
        return rng.choice(female_pool)
    return rng.choice(male_pool)


def relationship_for_member(age: int, member_index: int, host_age: int, rng: random.Random) -> str:
    if member_index == 1:
        return "Chủ hộ"
    if age >= host_age - 5 and age <= host_age + 5 and member_index == 2:
        return rng.choice(["Vợ/chồng", "Vợ/chồng chủ hộ"])
    if age < 18:
        return rng.choice(["Con", "Cháu"])
    if age >= host_age + 18:
        return rng.choice(["Bố/mẹ", "Khác"])
    return rng.choice(["Con", "Anh/chị/em", "Khác"])


def choose_child_health_insurance(is_poor: bool, is_near_poor: bool, rng: random.Random) -> bool:
    if is_poor:
        return rng.random() < 0.70
    if is_near_poor:
        return rng.random() < 0.78
    return rng.random() < 0.88


def choose_child_deprivation_flag(base_prob: float, rng: random.Random) -> bool:
    return rng.random() < base_prob


def household_deprivation_flags(classify: str | None, rng: random.Random, target_total: int | None = None) -> dict[str, bool]:
    poor = classify == "Hộ nghèo"
    near_poor = classify == "Hộ cận nghèo"
    severity_map = {
        "poor": (3, 8, [0.05, 0.08, 0.10, 0.14, 0.16, 0.17, 0.16, 0.14]),
        "near_poor": (0, 2, [0.35, 0.45, 0.20]),
        "other": (0, 2, [0.55, 0.30, 0.15]),
    }
    bucket = classify_bucket(classify)
    low, high, weights = severity_map[bucket]
    if target_total is None:
        target_total = choose_from_weights(rng, list(range(low, high + 1)), weights)
    target_total = int(max(low, min(high, target_total)))
    all_flags = [
        "employment",
        "dependentPerson",
        "nutrition",
        "healthInsurance",
        "adultEducation",
        "childSchoolAttendance",
        "housingQuality",
        "housingArea",
        "cleanWater",
        "hygienicToilet",
        "telecommunication",
        "informationAccessAssets",
    ]
    weights_map = {
        "employment": 1.0,
        "dependentPerson": 0.8,
        "nutrition": 1.1,
        "healthInsurance": 1.2,
        "adultEducation": 0.7,
        "childSchoolAttendance": 1.0,
        "housingQuality": 1.0,
        "housingArea": 0.9,
        "cleanWater": 0.8,
        "hygienicToilet": 0.85,
        "telecommunication": 0.6,
        "informationAccessAssets": 0.75,
    }
    if target_total <= 0:
        chosen = set()
    else:
        ordered = sorted(all_flags, key=lambda k: (weights_map[k], rng.random()), reverse=True)
        chosen = set(ordered[:target_total])
    flags = {f"deprivation.{name}": (name in chosen) for name in all_flags}
    if poor and not any(flags.values()):
        flags["deprivation.healthInsurance"] = True
    if near_poor and not any(flags.values()) and target_total > 0:
        flags["deprivation.healthInsurance"] = True
    return flags


def reason_flags(classify: str | None, deprivation_flags: dict[str, bool], children_total: int, rng: random.Random) -> dict[str, bool]:
    poor = classify == "Hộ nghèo"
    near_poor = classify == "Hộ cận nghèo"
    if not (poor or near_poor):
        return {
            "reason.lackProductionLand": False,
            "reason.lackCapital": False,
            "reason.lackLabor": False,
            "reason.lackProductionTools": False,
            "reason.lackProductionKnowledge": False,
            "reason.lackLaborSkill": False,
            "reason.illnessOrAccident": False,
            "reason.other": False,
        }
    base = 0.25 if poor else 0.18
    choices = {
        "reason.lackProductionLand": 0.10 + base,
        "reason.lackCapital": 0.18 + base,
        "reason.lackLabor": 0.12 + base,
        "reason.lackProductionTools": 0.10 + base,
        "reason.lackProductionKnowledge": 0.10 + base,
        "reason.lackLaborSkill": 0.10 + base,
        "reason.illnessOrAccident": 0.14 + base,
        "reason.other": 0.03,
    }
    if deprivation_flags.get("deprivation.healthInsurance"):
        choices["reason.illnessOrAccident"] += 0.08
    if deprivation_flags.get("deprivation.housingQuality") or deprivation_flags.get("deprivation.housingArea"):
        choices["reason.lackCapital"] += 0.04
    if children_total > 0:
        choices["reason.lackCapital"] += 0.02
    flags = {k: rng.random() < p for k, p in choices.items()}
    if not any(flags.values()):
        flags["reason.other"] = True
    return flags


def support_flags(
    classify: str | None,
    deprivation_flags: dict[str, bool],
    reason_flags_map: dict[str, bool],
    children_total: int,
    children_lack_health: int,
    rng: random.Random,
) -> dict[str, bool]:
    poor = classify == "Hộ nghèo"
    near_poor = classify == "Hộ cận nghèo"
    if not (poor or near_poor):
        return {
            "support.health": False,
            "support.education": False,
            "support.production": False,
            "support.credit": False,
            "support.housing": False,
            "support.other": False,
        }
    flags = {
        "support.health": bool(deprivation_flags.get("deprivation.healthInsurance") or children_lack_health > 0),
        "support.education": bool(children_total > 0 and (deprivation_flags.get("deprivation.childSchoolAttendance") or children_total > 0)),
        "support.production": bool(
            reason_flags_map.get("reason.lackProductionLand")
            or reason_flags_map.get("reason.lackProductionTools")
            or reason_flags_map.get("reason.lackProductionKnowledge")
        ),
        "support.credit": bool(reason_flags_map.get("reason.lackCapital")),
        "support.housing": bool(deprivation_flags.get("deprivation.housingQuality") or deprivation_flags.get("deprivation.housingArea")),
        "support.other": False,
    }
    if not any(v for k, v in flags.items() if k != "support.other"):
        flags["support.other"] = True
    return flags


def poverty_detail(classify: str | None, rng: random.Random) -> tuple[str | None, str | None]:
    if classify == "Hộ nghèo":
        return choose_from_weights(rng, ["Nghèo mới", "Tái nghèo", "Nghèo cũ"], [0.55, 0.25, 0.20]), "Không áp dụng"
    if classify == "Hộ cận nghèo":
        return "Không áp dụng", choose_from_weights(rng, ["Cận nghèo mới", "Tái cận nghèo", "Cận nghèo cũ"], [0.55, 0.20, 0.25])
    return "Không áp dụng", "Không áp dụng"


def class_transition_label(
    beginning: str | None,
    ending: str | None,
    kind: str,
    rng: random.Random,
) -> tuple[str, str]:
    poor_label, near_label, _ = resolve_transition_change_types(beginning or "Hộ không nghèo", ending or "Hộ không nghèo")
    return poor_label, near_label


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    for col in df.columns:
        df[col] = df[col].map(clean_text)
    return df


def remove_header_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0
    mask = df.apply(is_header_row, axis=1)
    removed = int(mask.sum())
    return df.loc[~mask].reset_index(drop=True), removed


def normalize_raw_core_values(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["family.hostName", "family.coDanTocTaiCho", "family.hostGender", "quickReview.result", "classify", "reviewer"]:
        if col in df.columns:
            df[col] = df[col].map(clean_text)
    if "family.code" in df.columns:
        df["family.code"] = df["family.code"].map(clean_code)
    if "family.numberOfMembers" in df.columns:
        df["family.numberOfMembers"] = df["family.numberOfMembers"].map(lambda x: parse_int(x, default=1))
    if "b1Point" in df.columns:
        df["b1Point"] = df["b1Point"].map(lambda x: clean_text(x))
    if "b2Point" in df.columns:
        df["b2Point"] = df["b2Point"].map(lambda x: clean_text(x))
    return df


def deduplicate_core_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    dup_mask = df.duplicated(subset=CORE_COLUMNS, keep="first")
    removed = int(dup_mask.sum())
    return df.loc[~dup_mask].reset_index(drop=True), removed


def fix_duplicate_family_codes(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    df["processing.original_family_code"] = df["family.code"]
    df["processing.family_code_was_changed"] = False
    df["processing.duplicate_code_group_size"] = 1
    df["processing.note"] = ""
    changed = 0
    grouped = df.groupby("processing.original_family_code", dropna=False, sort=False)
    for code, idxs in grouped.indices.items():
        idx_list = list(idxs)
        if len(idx_list) <= 1:
            continue
        group_size = len(idx_list)
        df.loc[idx_list, "processing.duplicate_code_group_size"] = group_size
        for suffix_idx, pos in enumerate(idx_list[1:], start=1):
            new_code = f"{code}_DUP{suffix_idx:03d}"
            df.at[pos, "family.code"] = new_code
            df.at[pos, "processing.family_code_was_changed"] = True
            df.at[pos, "processing.note"] = "duplicate_family_code_suffix_applied"
            changed += 1
        df.at[idx_list[0], "processing.note"] = "duplicate_family_code_group_kept"
    return df, changed


def normalize_dates_for_year(
    df: pd.DataFrame,
    year: int,
    district: str,
    seed: int,
) -> pd.DataFrame:
    df = df.copy()
    original = df["date"].copy()
    normalized: list[str] = []
    changed_flags: list[bool] = []
    filled_flags: list[bool] = []
    sources: list[str] = []
    for value in original.tolist():
        parsed = parse_date_value(value)
        row_idx = len(normalized)
        row_rng = rng_for(seed, "date", year, district, row_idx, df.iloc[row_idx].get("family.code"))
        candidate_dates = candidate_dates_for_year(year)
        if parsed is not None and parsed.year == year:
            new_date = format_date(parsed)
            changed = format_date(parsed) != clean_text(value)
            filled = False
            source = "cleaned_original"
        else:
            new_date = row_rng.choice(candidate_dates)
            changed = True
            filled = True
            source = "filled_from_legal_period_rule"
        normalized.append(new_date)
        changed_flags.append(changed)
        filled_flags.append(filled)
        sources.append(source)
    df["processing.original_date"] = original.map(clean_text)
    df["processing.source.date"] = sources
    df["processing.date_was_filled"] = filled_flags
    df["date"] = normalized
    df["processing.date_was_normalized"] = changed_flags
    df["administrative.year"] = year
    return df


def normalize_co_dan_toc(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["family.coDanTocTaiCho"] = df["family.coDanTocTaiCho"].map(normalize_yes_no)
    df["family.isDTTC"] = df["family.coDanTocTaiCho"].eq("Có")
    return df


def normalize_host_gender(df: pd.DataFrame, district: str, year: int, seed: int) -> pd.DataFrame:
    df = df.copy()
    genders = []
    for idx, row in df.iterrows():
        rng = rng_for(seed, "gender", year, district, idx, row.get("family.code"), row.get("family.hostName"))
        genders.append(normalize_gender(row.get("family.hostGender"), row.get("family.hostName"), rng))
    df["family.hostGender"] = genders
    return df


def assign_hierarchy(df: pd.DataFrame, year: int, district: str, seed: int, commune_mapping: dict[str, list[str]] | None = None) -> pd.DataFrame:
    df = df.copy()
    communes = resolve_district_communes(district, commune_mapping)
    rng = rng_for(seed, "communes", year, district, len(df))
    df["administrative.province"] = "Đắk Nông"
    df["administrative.district"] = district
    df["administrative.commune"] = assign_communes(communes, len(df), rng)
    df["administrative.village_or_group"] = [assign_village_or_group(rng_for(seed, "village", year, district, i, code)) for i, code in enumerate(df["family.code"].tolist(), start=1)]
    return df


def generate_household_features(df: pd.DataFrame, year: int, district: str) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    df = df.copy()
    members_rows: list[dict[str, Any]] = []

    poverty_details = []
    near_poverty_details = []
    host_birth_dates = []
    host_birth_years = []
    deprivation_total_counts = []
    children_total_counts = []
    children_lack_health_counts = []
    children_nutrition_counts = []
    children_school_counts = []
    is_medium_living = []
    has_no_labor_capacity = []
    has_merit_policy = []
    ethnicity_values = []
    is_dtts_values = []
    is_kinh_values = []
    support_health_values = []
    support_education_values = []
    support_production_values = []
    support_credit_values = []
    support_housing_values = []
    support_other_values = []
    reason_columns = {f"reason.{name}": [] for name in [
        "lackProductionLand",
        "lackCapital",
        "lackLabor",
        "lackProductionTools",
        "lackProductionKnowledge",
        "lackLaborSkill",
        "illnessOrAccident",
        "other",
    ]}
    deprivation_columns = {f"deprivation.{name}": [] for name in [
        "employment",
        "dependentPerson",
        "nutrition",
        "healthInsurance",
        "adultEducation",
        "childSchoolAttendance",
        "housingQuality",
        "housingArea",
        "cleanWater",
        "hygienicToilet",
        "telecommunication",
        "informationAccessAssets",
    ]}
    transition_beginning = []
    transition_ending = []
    poor_change_types = []
    near_change_types = []
    members_generated_flags = []
    members_file_paths = []
    members_json_values = []

    household_order = []

    for idx, row in df.iterrows():
        household_order.append(idx + 1)
        rng = rng_for("household", year, district, idx, row.get("family.code"), row.get("family.hostName"))
        host_age = host_age_for_row(row, rng)
        host_birth_date = age_to_birthdate(host_age, rng, year)
        host_birth_dates.append(format_date(host_birth_date))
        host_birth_years.append(host_birth_date.year)

        ethnicity = choose_ethnicity(district, bool(row.get("family.isDTTC")), rng)
        if ethnicity == "Kinh" and row.get("family.isDTTC"):
            ethnicity = choose_ethnicity(district, True, rng)
        ethnicity_values.append(ethnicity)
        is_kinh = ethnicity == "Kinh"
        is_dtts = not is_kinh
        if is_kinh:
            row_is_dttc = False
        else:
            row_is_dttc = True if row.get("family.isDTTC") else False
        is_dtts_values.append(is_dtts)
        is_kinh_values.append(is_kinh)
        df.at[idx, "family.isDTTC"] = bool(row_is_dttc)

        poor_detail, near_detail = poverty_detail(row.get("classify"), rng)
        poverty_details.append(poor_detail)
        near_poverty_details.append(near_detail)

        no_labor_prob = 0.18 if row.get("classify") == "Hộ nghèo" else 0.12 if row.get("classify") == "Hộ cận nghèo" else 0.04
        if host_age >= 70:
            no_labor_prob += 0.10
        if parse_int(row.get("family.numberOfMembers"), 1) == 1:
            no_labor_prob += 0.06
        has_no_labor_capacity.append(rng.random() < no_labor_prob)

        has_merit_policy.append(rng.random() < 0.03)

        size = parse_int(row.get("family.numberOfMembers"), default=1)
        ages = allocate_member_ages(size, row.get("classify"), host_age, rng)
        member_count = len(ages)
        child_indices = [i for i, age in enumerate(ages) if age < 16]
        children_total = len(child_indices)

        # household-level deprivation flags
        dep_flags = household_deprivation_flags(row.get("classify"), rng)
        # Make child-related indicators align with child members later
        if children_total == 0:
            dep_flags["deprivation.childSchoolAttendance"] = False
        if row.get("classify") == "Hộ không nghèo" and rng.random() < 0.6:
            dep_flags["deprivation.healthInsurance"] = False

        # Member generation first so household child counts can be derived
        member_records: list[dict[str, Any]] = []
        child_lack_health = 0
        child_nutrition = 0
        child_school = 0
        for member_pos, age in enumerate(ages, start=1):
            member_rng = rng_for("member", year, district, idx, member_pos, row.get("family.code"), row.get("family.hostName"))
            gender = row.get("family.hostGender") if member_pos == 1 else ("Nam" if member_rng.random() < 0.5 else "Nữ")
            if member_pos == 1:
                full_name = row.get("family.hostName")
                relation = "Chủ hộ"
                ethnicity_member = ethnicity
                birth_date = host_birth_date
                birth_year = host_birth_date.year
                is_host = True
            else:
                full_name = generate_member_name(member_rng, gender, ethnicity)
                relation = relationship_for_member(age, member_pos, host_age, member_rng)
                ethnicity_member = ethnicity if member_rng.random() < 0.92 else choose_ethnicity(district, False, member_rng)
                birth_date = age_to_birthdate(age, member_rng, year)
                birth_year = birth_date.year
                is_host = False
            is_child = age < 16
            if is_child:
                lack_health = not choose_child_health_insurance(row.get("classify") == "Hộ nghèo", row.get("classify") == "Hộ cận nghèo", member_rng)
                nutrition_deprived = choose_child_deprivation_flag(0.28 if row.get("classify") == "Hộ nghèo" else 0.18 if row.get("classify") == "Hộ cận nghèo" else 0.08, member_rng)
                school_deprived = choose_child_deprivation_flag(0.24 if row.get("classify") == "Hộ nghèo" else 0.14 if row.get("classify") == "Hộ cận nghèo" else 0.05, member_rng)
            else:
                lack_health = False
                nutrition_deprived = False
                school_deprived = False
            child_lack_health += int(is_child and lack_health)
            child_nutrition += int(is_child and nutrition_deprived)
            child_school += int(is_child and school_deprived)
            member_records.append(
                {
                    "administrative.year": year,
                    "administrative.province": "Đắk Nông",
                    "administrative.district": district,
                    "administrative.commune": row.get("administrative.commune"),
                    "administrative.village_or_group": row.get("administrative.village_or_group"),
                    "family.code": row.get("family.code"),
                    "family.hostName": row.get("family.hostName"),
                    "member.householdOrder": idx + 1,
                    "member.orderInHousehold": member_pos,
                    "member.fullName": full_name,
                    "member.gender": gender,
                    "member.birthDate": format_date(birth_date),
                    "member.birthYear": birth_year,
                    "member.ethnicity": ethnicity_member,
                    "member.relationshipToHost": relation,
                    "member.isHost": is_host,
                    "member.isChild": is_child,
                    "member.hasHealthInsurance": bool(not lack_health),
                    "member.nutritionDeprived": bool(nutrition_deprived),
                    "member.schoolAttendanceDeprived": bool(school_deprived),
                }
            )

        children_total_counts.append(children_total)
        children_lack_health_counts.append(min(children_total, child_lack_health))
        children_nutrition_counts.append(min(children_total, child_nutrition))
        children_school_counts.append(min(children_total, child_school))
        deprivation_total = int(sum(bool(v) for v in dep_flags.values()))
        deprivation_total_counts.append(deprivation_total)
        for col, val in dep_flags.items():
            deprivation_columns[col].append(bool(val))

        # Transition labels
        if year == 2023:
            beginning = row.get("classify")
        else:
            beginning = None
        transition_beginning.append(beginning)
        transition_ending.append(row.get("classify"))
        poor_label, near_label = class_transition_label(beginning, row.get("classify"), "poor" if row.get("classify") == "Hộ nghèo" else "near" if row.get("classify") == "Hộ cận nghèo" else "other", rng)
        poor_change_types.append(poor_label)
        near_change_types.append(near_label)

        # Reasons
        reasons = reason_flags(row.get("classify"), dep_flags, children_total, rng)
        for col, val in reasons.items():
            reason_columns[col].append(bool(val))

        # Supports depend on reasons and deprivations
        supports = support_flags(row.get("classify"), dep_flags, reasons, children_total, child_lack_health, rng)
        support_health_values.append(supports["support.health"])
        support_education_values.append(supports["support.education"])
        support_production_values.append(supports["support.production"])
        support_credit_values.append(supports["support.credit"])
        support_housing_values.append(supports["support.housing"])
        support_other_values.append(supports["support.other"])

        # medium living standard flag
        if row.get("classify") == "Hộ không nghèo":
            is_medium_living.append(rng.random() < 0.28)
        elif row.get("classify") == "Hộ cận nghèo":
            is_medium_living.append(rng.random() < 0.04)
        else:
            is_medium_living.append(rng.random() < 0.01)

        members_generated_flags.append(True)
        members_file_paths.append("")
        members_json_values.append(json.dumps(member_records, ensure_ascii=False))
        members_rows.extend(member_records)

    df["family.hostBirthDate"] = host_birth_dates
    df["family.hostBirthYear"] = host_birth_years
    df["family.ethnicity"] = ethnicity_values
    df["family.isDTTS"] = is_dtts_values
    df["family.isKinh"] = is_kinh_values
    df["family.hasNoLaborCapacity"] = has_no_labor_capacity
    df["family.hasRevolutionMeritPolicy"] = has_merit_policy
    df["family.povertyStatusDetail"] = poverty_details
    df["family.nearPovertyStatusDetail"] = near_poverty_details
    df["family.isAgricultureForestryFisherySaltMediumLivingStandard"] = is_medium_living
    df["deprivation.totalCount"] = deprivation_total_counts
    df["children.totalCount"] = children_total_counts
    df["children.lackHealthInsuranceCount"] = children_lack_health_counts
    df["children.nutritionDeprivedCount"] = children_nutrition_counts
    df["children.schoolAttendanceDeprivedCount"] = children_school_counts
    df["transition.beginningClassify"] = transition_beginning
    df["transition.endingClassify"] = transition_ending
    df["transition.poorChangeType"] = poor_change_types
    df["transition.nearPoorChangeType"] = near_change_types
    df["family.membersGenerated"] = members_generated_flags
    df["family.membersFile"] = members_file_paths
    df["family.membersJson"] = members_json_values
    for col, values in deprivation_columns.items():
        df[col] = values
    for col, values in reason_columns.items():
        df[col] = values
    df["support.health"] = support_health_values
    df["support.education"] = support_education_values
    df["support.production"] = support_production_values
    df["support.credit"] = support_credit_values
    df["support.housing"] = support_housing_values
    df["support.other"] = support_other_values
    df["member.householdOrder"] = household_order
    return df, members_rows


def generate_household_features(
    df: pd.DataFrame,
    year: int,
    district: str,
    seed: int,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    df = df.copy()
    rows: list[dict[str, Any]] = []
    members_rows: list[dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_rng = rng_for(seed, "household", year, district, idx, row.get("family.code"), row.get("family.hostName"))
        out: dict[str, Any] = {col: row.get(col) for col in df.columns}

        out["processing.note"] = clean_text(row.get("processing.note")) or "Không áp dụng"
        out["processing.duplicate_core_removed"] = bool(row.get("processing.duplicate_core_removed"))
        out["processing.duplicate_code_group_size"] = parse_int(row.get("processing.duplicate_code_group_size"), default=1)

        original_classify = clean_text(row.get("classify")) or "Hộ không nghèo"
        if original_classify not in {"Hộ nghèo", "Hộ cận nghèo", "Hộ không nghèo"}:
            original_classify = "Hộ không nghèo"
        out["classify.original"] = original_classify

        area_type, area_source, area_confidence = resolve_area_type_info(row.get("administrative.commune"))
        out["administrative.areaType"] = area_type
        out["administrative.areaTypeSource"] = area_source
        out["administrative.areaTypeConfidence"] = area_confidence

        review_result, review_source, review_filled = resolve_quick_review_result(row.get("quickReview.result"))
        reviewer_value, reviewer_source, reviewer_filled = resolve_reviewer_value(row.get("reviewer"))
        out["quickReview.result"] = review_result
        out["reviewer"] = reviewer_value
        out["reviewer.role"] = "Người thực hiện rà soát"
        out["reviewer.source"] = reviewer_source

        date_value = row.get("date")
        out["review.reviewType"] = resolve_review_type(date_value)
        out["review.reviewPeriod"] = review_period_for_year(year)

        ethnicity, co_dan_toc, is_dttc = resolve_ethnicity_and_dttc(district, original_classify, row.get("family.coDanTocTaiCho"), row_rng)
        host_gender = normalize_gender(row.get("family.hostGender"), row.get("family.hostName"), row_rng)
        host_name, host_name_source, host_name_filled = resolve_host_name_value(row.get("family.hostName"), host_gender, ethnicity, row_rng)

        out["family.hostName"] = host_name
        out["family.hostGender"] = host_gender
        out["family.coDanTocTaiCho"] = co_dan_toc
        out["family.ethnicity"] = ethnicity
        out["family.isDTTC"] = bool(is_dttc)
        out["family.isKinh"] = ethnicity == "Kinh"
        out["family.isDTTS"] = not out["family.isKinh"]

        out["processing.original_family_hostName"] = clean_text(row.get("family.hostName")) or "Không xác định"
        out["processing.source.family.hostName"] = host_name_source
        out["processing.family_hostName_was_filled"] = host_name_filled
        out["processing.family_hostName_rule"] = "generated_by_controlled_synthetic_rule"
        out["processing.original_family_hostGender"] = clean_text(row.get("family.hostGender")) or "Không xác định"
        out["processing.source.family.hostGender"] = "cleaned_original" if not is_missingish(row.get("family.hostGender")) else "generated_by_controlled_synthetic_rule"
        out["processing.family_hostGender_was_filled"] = is_missingish(row.get("family.hostGender"))
        out["processing.family_hostGender_rule"] = "normalized_gender_rule"
        out["processing.original_family_coDanTocTaiCho"] = clean_text(row.get("family.coDanTocTaiCho")) or "Không xác định"
        out["processing.source.family.coDanTocTaiCho"] = "cleaned_original" if not is_missingish(row.get("family.coDanTocTaiCho")) else "generated_by_controlled_synthetic_rule"
        out["processing.family_coDanTocTaiCho_was_filled"] = is_missingish(row.get("family.coDanTocTaiCho"))
        out["processing.family_coDanTocTaiCho_rule"] = "ethnicity_consistency_rule"
        out["family.isPoorNearPoorDTTS"] = bool(original_classify in {"Hộ nghèo", "Hộ cận nghèo"} and out["family.isDTTS"])

        family_size, family_size_source, family_size_filled = resolve_family_members_count(row.get("family.numberOfMembers"), original_classify, row_rng)
        out["family.numberOfMembers"] = int(family_size)
        out["processing.original_family_numberOfMembers"] = parse_int(row.get("family.numberOfMembers"), default=0) if not is_missingish(row.get("family.numberOfMembers")) else "Không xác định"
        out["processing.source.family.numberOfMembers"] = family_size_source
        out["processing.family_numberOfMembers_was_filled"] = family_size_filled
        out["processing.family_numberOfMembers_rule"] = "generated_from_classify_final_when_invalid"

        b1_point, b1_source, b1_filled = resolve_b1_point(original_classify, area_type, row.get("b1Point"), row_rng)
        deprivation_total, b2_point, b2_source, b2_filled = resolve_deprivation_total_count(original_classify, row.get("b2Point"), row.get("deprivation.totalCount"), row_rng)
        legal_classify = resolve_classification_from_scores(b1_point, deprivation_total, area_type)
        if original_classify == legal_classify:
            final_classify = original_classify
            consistency_status = "matched_legal_rule"
            consistency_reason = "original_consistent_with_legal_rule"
        else:
            final_classify = legal_classify
            consistency_status = "overridden_by_legal_rule" if original_classify in {"Hộ nghèo", "Hộ cận nghèo", "Hộ không nghèo"} else "filled_from_legal_rule"
            consistency_reason = "original_missing" if original_classify not in {"Hộ nghèo", "Hộ cận nghèo", "Hộ không nghèo"} else "original_conflicted_with_legal_rule"

        out["classify"] = original_classify
        out["classify.legalRecomputed"] = legal_classify
        out["classify.final"] = final_classify
        out["classify.consistencyStatus"] = consistency_status
        out["classify.consistencyReason"] = consistency_reason
        out["b1Point"] = int(b1_point)
        out["b2Point"] = int(b2_point)
        out["deprivation.totalCount"] = int(deprivation_total)
        out["processing.original_b1Point"] = clean_text(row.get("b1Point")) or "Không xác định"
        out["processing.source.b1Point"] = b1_source
        out["processing.b1Point_was_filled"] = b1_filled
        out["processing.b1Point_rule"] = f"{area_type}_threshold_rule"
        out["processing.original_b2Point"] = clean_text(row.get("b2Point")) or "Không xác định"
        out["processing.source.b2Point"] = b2_source
        out["processing.source.deprivation.totalCount"] = "derived_from_generated_values"
        out["processing.b2Point_was_filled"] = b2_filled
        out["processing.b2Point_rule"] = "deprivation.totalCount * 10"

        host_age = host_age_for_row(pd.Series({"classify": final_classify, "family.numberOfMembers": family_size}), row_rng)
        host_birth_date = age_to_birthdate(host_age, row_rng, year)
        out["family.hostBirthDate"] = format_date(host_birth_date)
        out["family.hostBirthYear"] = host_birth_date.year
        out["processing.original_date"] = clean_text(row.get("processing.original_date")) or clean_text(row.get("date")) or "Không xác định"
        out["processing.source.date"] = clean_text(row.get("processing.source.date")) or "cleaned_original"
        out["processing.date_was_filled"] = bool(row.get("processing.date_was_filled")) if row.get("processing.date_was_filled") is not None else False
        out["processing.date_was_normalized"] = bool(row.get("processing.date_was_normalized")) if row.get("processing.date_was_normalized") is not None else False

        if final_classify == "Hộ nghèo":
            out["family.povertyStatusDetail"], out["family.nearPovertyStatusDetail"] = poverty_detail(final_classify, row_rng)
        elif final_classify == "Hộ cận nghèo":
            out["family.povertyStatusDetail"], out["family.nearPovertyStatusDetail"] = poverty_detail(final_classify, row_rng)
        else:
            out["family.povertyStatusDetail"], out["family.nearPovertyStatusDetail"] = "Không áp dụng", "Không áp dụng"
        out["family.isAgricultureForestryFisherySaltMediumLivingStandard"] = bool(final_classify == "Hộ không nghèo" and row_rng.random() < 0.28)
        out["family.hasNoLaborCapacity"] = bool(
            (final_classify == "Hộ nghèo" and row_rng.random() < 0.18)
            or (final_classify == "Hộ cận nghèo" and row_rng.random() < 0.12)
            or (final_classify == "Hộ không nghèo" and row_rng.random() < 0.04)
            or host_age >= 70
        )
        out["family.hasRevolutionMeritPolicy"] = bool(row_rng.random() < 0.03)

        size = int(family_size)
        ages = allocate_member_ages(size, final_classify, host_age, row_rng)
        child_indices = [i for i, age in enumerate(ages) if age < 16]
        child_total = len(child_indices)
        child_lack_health = 0
        child_nutrition = 0
        child_school = 0
        member_records: list[dict[str, Any]] = []
        for member_pos, age in enumerate(ages, start=1):
            member_rng = rng_for(seed, "member", year, district, idx, member_pos, row.get("family.code"), host_name)
            if member_pos == 1:
                member_name = host_name
                relation = "Chủ hộ"
                member_gender = out["family.hostGender"]
                member_ethnicity = ethnicity
                birth_date = host_birth_date
                birth_year = host_birth_date.year
                is_host = True
            else:
                member_gender = "Nam" if member_rng.random() < 0.5 else "Nữ"
                member_name = generate_member_name(member_rng, member_gender, ethnicity)
                relation = relationship_for_member(age, member_pos, host_age, member_rng)
                member_ethnicity = ethnicity if member_rng.random() < 0.92 else choose_ethnicity(district, False, member_rng)
                birth_date = age_to_birthdate(age, member_rng, year)
                birth_year = birth_date.year
                is_host = False
            is_child = age < 16
            if is_child:
                lack_health = not choose_child_health_insurance(final_classify == "Hộ nghèo", final_classify == "Hộ cận nghèo", member_rng)
                nutrition_deprived = choose_child_deprivation_flag(0.28 if final_classify == "Hộ nghèo" else 0.18 if final_classify == "Hộ cận nghèo" else 0.08, member_rng)
                school_deprived = choose_child_deprivation_flag(0.24 if final_classify == "Hộ nghèo" else 0.14 if final_classify == "Hộ cận nghèo" else 0.05, member_rng)
            else:
                lack_health = False
                nutrition_deprived = False
                school_deprived = False
            child_lack_health += int(is_child and lack_health)
            child_nutrition += int(is_child and nutrition_deprived)
            child_school += int(is_child and school_deprived)
            member_records.append(
                {
                    "administrative.year": year,
                    "administrative.province": "Đắk Nông",
                    "administrative.district": district,
                    "administrative.commune": row.get("administrative.commune"),
                    "administrative.village_or_group": row.get("administrative.village_or_group"),
                    "family.code": out["family.code"],
                    "family.hostName": host_name,
                    "member.householdOrder": idx + 1,
                    "member.orderInHousehold": member_pos,
                    "member.fullName": member_name,
                    "member.gender": member_gender,
                    "member.birthDate": format_date(birth_date),
                    "member.birthYear": birth_year,
                    "member.ethnicity": member_ethnicity,
                    "member.relationshipToHost": relation,
                    "member.isHost": is_host,
                    "member.isChild": is_child,
                    "member.hasHealthInsurance": bool(not lack_health),
                    "member.nutritionDeprived": bool(nutrition_deprived),
                    "member.schoolAttendanceDeprived": bool(school_deprived),
                }
            )

        out["children.totalCount"] = child_total
        out["children.lackHealthInsuranceCount"] = min(child_total, child_lack_health)
        out["children.nutritionDeprivedCount"] = min(child_total, child_nutrition)
        out["children.schoolAttendanceDeprivedCount"] = min(child_total, child_school)

        dep_flags = household_deprivation_flags(final_classify, row_rng, target_total=deprivation_total)
        for col, val in dep_flags.items():
            out[col] = bool(val)

        reason_map = reason_flags(final_classify, dep_flags, child_total, row_rng)
        for col, val in reason_map.items():
            out[col] = bool(val)

        support_map = support_flags(final_classify, dep_flags, reason_map, child_total, child_lack_health, row_rng)
        for col, val in support_map.items():
            out[col] = bool(val)

        if final_classify == "Hộ nghèo":
            beginning_pool = ["Hộ nghèo", "Hộ cận nghèo", "Hộ không nghèo"]
        elif final_classify == "Hộ cận nghèo":
            beginning_pool = ["Hộ cận nghèo", "Hộ nghèo", "Hộ không nghèo"]
        else:
            beginning_pool = ["Hộ không nghèo", "Hộ cận nghèo", "Hộ nghèo"]
        beginning = row_rng.choice(beginning_pool)
        poor_change, near_change, escaped = resolve_transition_change_types(beginning, final_classify)
        out["transition.beginningClassify"] = beginning
        out["transition.endingClassify"] = final_classify
        out["transition.poorChangeType"] = poor_change
        out["transition.nearPoorChangeType"] = near_change
        out["transition.isEscapedPoverty"] = escaped

        out["family.membersGenerated"] = True
        out["family.membersFile"] = "Không áp dụng"
        out["family.membersJson"] = json.dumps(member_records, ensure_ascii=False)

        out["member.householdOrder"] = idx + 1

        out["processing.source.family.code"] = "derived_from_original" if out.get("processing.family_code_was_changed") else "original"
        out["processing.family_code_rule"] = "duplicate suffix preserved from first occurrence"

        rows.append(out)
        members_rows.extend(member_records)

    df = pd.DataFrame(rows)
    for col in [
        "family.coDanTocTaiCho",
        "family.hostGender",
        "family.hostName",
        "family.numberOfMembers",
        "quickReview.result",
        "b1Point",
        "b2Point",
        "classify",
        "reviewer",
        "reviewer.role",
        "reviewer.source",
        "review.reviewType",
        "review.reviewPeriod",
        "administrative.areaType",
        "administrative.areaTypeSource",
        "administrative.areaTypeConfidence",
        "classify.original",
        "classify.legalRecomputed",
        "classify.final",
        "classify.consistencyStatus",
        "classify.consistencyReason",
        "transition.isEscapedPoverty",
        "family.isPoorNearPoorDTTS",
        "processing.source.date",
        "processing.date_was_filled",
        "processing.date_was_normalized",
        "processing.original_b1Point",
        "processing.source.b1Point",
        "processing.b1Point_was_filled",
        "processing.b1Point_rule",
        "processing.original_b2Point",
        "processing.source.b2Point",
        "processing.source.deprivation.totalCount",
        "processing.b2Point_was_filled",
        "processing.b2Point_rule",
        "processing.original_family_hostName",
        "processing.source.family.hostName",
        "processing.family_hostName_was_filled",
        "processing.family_hostName_rule",
        "processing.original_family_hostGender",
        "processing.source.family.hostGender",
        "processing.family_hostGender_was_filled",
        "processing.family_hostGender_rule",
        "processing.original_family_coDanTocTaiCho",
        "processing.source.family.coDanTocTaiCho",
        "processing.family_coDanTocTaiCho_was_filled",
        "processing.family_coDanTocTaiCho_rule",
        "processing.original_family_numberOfMembers",
        "processing.source.family.numberOfMembers",
        "processing.family_numberOfMembers_was_filled",
        "processing.family_numberOfMembers_rule",
        "processing.source.family.code",
        "processing.family_code_rule",
    ]:
        if col not in df.columns:
            if col in {
                "family.isDTTC",
                "family.isDTTS",
                "family.isKinh",
                "family.hasNoLaborCapacity",
                "family.hasRevolutionMeritPolicy",
                "family.isAgricultureForestryFisherySaltMediumLivingStandard",
                "family.membersGenerated",
                "transition.isEscapedPoverty",
                "family.isPoorNearPoorDTTS",
                "processing.date_was_filled",
                "processing.date_was_normalized",
                "processing.b1Point_was_filled",
                "processing.b2Point_was_filled",
                "processing.family_hostName_was_filled",
                "processing.family_hostGender_was_filled",
                "processing.family_coDanTocTaiCho_was_filled",
                "processing.family_numberOfMembers_was_filled",
            }:
                df[col] = False
            elif col in {
                "family.numberOfMembers",
                "b1Point",
                "b2Point",
                "deprivation.totalCount",
                "children.totalCount",
                "children.lackHealthInsuranceCount",
                "children.nutritionDeprivedCount",
                "children.schoolAttendanceDeprivedCount",
                "family.hostBirthYear",
                "member.householdOrder",
            }:
                df[col] = 0
            elif col == "family.membersJson":
                df[col] = "[]"
            else:
                df[col] = "Không xác định"

    df["family.membersGenerated"] = df["family.membersGenerated"].fillna(True)
    df["family.membersFile"] = df["family.membersFile"].fillna("Không áp dụng")
    df["family.membersJson"] = df["family.membersJson"].fillna("[]")
    df["transition.isEscapedPoverty"] = df["transition.isEscapedPoverty"].fillna(False)
    df["family.isPoorNearPoorDTTS"] = df["family.isPoorNearPoorDTTS"].fillna(False)

    return df, members_rows


def append_processing_columns(df: pd.DataFrame) -> pd.DataFrame:
    new_cols_in_order = [
        "processing.original_family_code",
        "processing.family_code_was_changed",
        "processing.duplicate_core_removed",
        "processing.duplicate_code_group_size",
        "processing.note",
        "processing.original_date",
        "processing.date_was_normalized",
        "administrative.year",
        "administrative.province",
        "administrative.district",
        "administrative.commune",
        "administrative.village_or_group",
        "family.isDTTC",
        "family.hostBirthDate",
        "family.hostBirthYear",
        "family.ethnicity",
        "family.isDTTS",
        "family.isKinh",
        "family.hasNoLaborCapacity",
        "family.hasRevolutionMeritPolicy",
        "family.povertyStatusDetail",
        "family.nearPovertyStatusDetail",
        "family.isAgricultureForestryFisherySaltMediumLivingStandard",
        "deprivation.totalCount",
        "children.totalCount",
        "children.lackHealthInsuranceCount",
        "children.nutritionDeprivedCount",
        "children.schoolAttendanceDeprivedCount",
        "transition.beginningClassify",
        "transition.endingClassify",
        "transition.poorChangeType",
        "transition.nearPoorChangeType",
        "family.membersGenerated",
        "family.membersFile",
        "family.membersJson",
        "support.health",
        "support.education",
        "support.production",
        "support.credit",
        "support.housing",
        "support.other",
        "reason.lackProductionLand",
        "reason.lackCapital",
        "reason.lackLabor",
        "reason.lackProductionTools",
        "reason.lackProductionKnowledge",
        "reason.lackLaborSkill",
        "reason.illnessOrAccident",
        "reason.other",
        "deprivation.employment",
        "deprivation.dependentPerson",
        "deprivation.nutrition",
        "deprivation.healthInsurance",
        "deprivation.adultEducation",
        "deprivation.childSchoolAttendance",
        "deprivation.housingQuality",
        "deprivation.housingArea",
        "deprivation.cleanWater",
        "deprivation.hygienicToilet",
        "deprivation.telecommunication",
        "deprivation.informationAccessAssets",
    ]
    for col in new_cols_in_order:
        if col not in df.columns:
            df[col] = None
    ordered = [c for c in df.columns if c not in new_cols_in_order] + new_cols_in_order
    return df[ordered]


def write_dataframe_to_excel(df: pd.DataFrame, path: Path, sheet_name: str) -> None:
    ensure_dir(path.parent)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


def write_validation_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    ensure_dir(path.parent)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)


def build_required_columns_by_report() -> dict[str, Any]:
    result = {}
    for spec in REPORT_SPECS:
        result[str(spec["report_id"])] = {
            "report_id": spec["report_id"],
            "report_name": spec["report_name"],
            "required_columns": spec["required_columns"],
            "notes": spec["notes"],
        }
    return result


def build_report_schema_summary(template_status: str = "missing_in_workspace") -> list[dict[str, Any]]:
    summary = []
    for spec in REPORT_SPECS:
        summary.append(
            {
                "report_id": spec["report_id"],
                "report_name": spec["report_name"],
                "sheet_name": spec["sheet_name"],
                "used_range": spec["used_range"],
                "header_rows": spec["header_rows"],
                "main_groups": spec["main_groups"],
                "required_columns": spec["required_columns"],
                "column_notes": spec["notes"],
                "template_status": template_status,
            }
        )
    return summary


def build_data_dictionary() -> dict[str, Any]:
    dictionary = {
        "date": {"type": "string", "scope": "original", "description": "Ngày rà soát theo định dạng dd/mm/yyyy"},
        "family.hostName": {"type": "string", "scope": "original", "description": "Chủ hộ"},
        "family.code": {"type": "string", "scope": "original", "description": "Mã hộ sau khi xử lý duplicate"},
        "family.coDanTocTaiCho": {"type": "string", "scope": "original", "description": "Có/Không về dân tộc tại chỗ"},
        "family.hostGender": {"type": "string", "scope": "original", "description": "Giới tính chủ hộ"},
        "family.numberOfMembers": {"type": "integer", "scope": "original", "description": "Số nhân khẩu"},
        "quickReview.result": {"type": "string", "scope": "original", "description": "Kết quả phiếu rà soát nhanh"},
        "b1Point": {"type": "string", "scope": "original", "description": "Điểm phiếu B1"},
        "b2Point": {"type": "string", "scope": "original", "description": "Điểm phiếu B2"},
        "classify": {"type": "string", "scope": "original", "description": "Phân loại hộ"},
        "reviewer": {"type": "string", "scope": "original", "description": "Người rà soát"},
        "processing.original_family_code": {"type": "string", "scope": "generated", "description": "Mã hộ gốc trước khi thêm hậu tố"},
        "processing.family_code_was_changed": {"type": "boolean", "scope": "generated", "description": "Mã hộ có được đổi hậu tố hay không"},
        "processing.duplicate_core_removed": {"type": "boolean", "scope": "derived", "description": "Dòng thuộc nhóm có duplicate lõi đã bị loại bớt"},
        "processing.duplicate_code_group_size": {"type": "integer", "scope": "derived", "description": "Kích thước nhóm mã hộ trùng"},
        "processing.note": {"type": "string", "scope": "generated", "description": "Ghi chú xử lý"},
        "processing.original_date": {"type": "string", "scope": "generated", "description": "Giá trị date gốc"},
        "processing.date_was_normalized": {"type": "boolean", "scope": "generated", "description": "Đã chuẩn hóa ngày hay chưa"},
        "administrative.year": {"type": "integer", "scope": "derived", "description": "Năm dữ liệu"},
        "administrative.province": {"type": "string", "scope": "generated", "description": "Tỉnh"},
        "administrative.district": {"type": "string", "scope": "generated", "description": "Huyện/thành phố"},
        "administrative.commune": {"type": "string", "scope": "generated", "description": "Xã/phường/thị trấn"},
        "administrative.village_or_group": {"type": "string", "scope": "generated", "description": "Thôn/bon/tổ dân phố"},
        "family.isDTTC": {"type": "boolean", "scope": "derived", "description": "Có dân tộc tại chỗ"},
        "family.hostBirthDate": {"type": "string", "scope": "generated", "description": "Ngày sinh chủ hộ"},
        "family.hostBirthYear": {"type": "integer", "scope": "generated", "description": "Năm sinh chủ hộ"},
        "family.ethnicity": {"type": "string", "scope": "generated", "description": "Dân tộc chủ hộ"},
        "family.isDTTS": {"type": "boolean", "scope": "derived", "description": "Hộ DTTS"},
        "family.isKinh": {"type": "boolean", "scope": "derived", "description": "Hộ Kinh"},
        "family.hasNoLaborCapacity": {"type": "boolean", "scope": "generated", "description": "Hộ không có khả năng lao động"},
        "family.hasRevolutionMeritPolicy": {"type": "boolean", "scope": "generated", "description": "Hộ có chính sách người có công"},
        "family.povertyStatusDetail": {"type": "string", "scope": "generated", "description": "Trạng thái nghèo chi tiết"},
        "family.nearPovertyStatusDetail": {"type": "string", "scope": "generated", "description": "Trạng thái cận nghèo chi tiết"},
        "family.isAgricultureForestryFisherySaltMediumLivingStandard": {"type": "boolean", "scope": "generated", "description": "Hộ có mức sống trung bình"},
        "deprivation.totalCount": {"type": "integer", "scope": "derived", "description": "Tổng số chỉ số thiếu hụt"},
        "children.totalCount": {"type": "integer", "scope": "derived", "description": "Tổng số trẻ em"},
        "children.lackHealthInsuranceCount": {"type": "integer", "scope": "derived", "description": "Số trẻ thiếu BHYT"},
        "children.nutritionDeprivedCount": {"type": "integer", "scope": "derived", "description": "Số trẻ thiếu dinh dưỡng"},
        "children.schoolAttendanceDeprivedCount": {"type": "integer", "scope": "derived", "description": "Số trẻ thiếu đi học"},
        "transition.beginningClassify": {"type": "string", "scope": "generated", "description": "Phân loại đầu kỳ"},
        "transition.endingClassify": {"type": "string", "scope": "generated", "description": "Phân loại cuối kỳ"},
        "transition.poorChangeType": {"type": "string", "scope": "generated", "description": "Kiểu biến động hộ nghèo"},
        "transition.nearPoorChangeType": {"type": "string", "scope": "generated", "description": "Kiểu biến động hộ cận nghèo"},
        "family.membersGenerated": {"type": "boolean", "scope": "generated", "description": "Đã sinh member file"},
        "family.membersFile": {"type": "string", "scope": "generated", "description": "Đường dẫn file member"},
        "family.membersJson": {"type": "string", "scope": "generated", "description": "JSON danh sách thành viên"},
        "support.health": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ y tế"},
        "support.education": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ giáo dục"},
        "support.production": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ sản xuất"},
        "support.credit": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ tín dụng"},
        "support.housing": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ nhà ở"},
        "support.other": {"type": "boolean", "scope": "generated", "description": "Hỗ trợ khác"},
    }
    for name in [
        "employment",
        "dependentPerson",
        "nutrition",
        "healthInsurance",
        "adultEducation",
        "childSchoolAttendance",
        "housingQuality",
        "housingArea",
        "cleanWater",
        "hygienicToilet",
        "telecommunication",
        "informationAccessAssets",
    ]:
        dictionary[f"deprivation.{name}"] = {"type": "boolean", "scope": "generated", "description": "Chỉ số thiếu hụt"}
    for name in [
        "lackProductionLand",
        "lackCapital",
        "lackLabor",
        "lackProductionTools",
        "lackProductionKnowledge",
        "lackLaborSkill",
        "illnessOrAccident",
        "other",
    ]:
        dictionary[f"reason.{name}"] = {"type": "boolean", "scope": "generated", "description": "Nguyên nhân nghèo/cận nghèo"}
    return dictionary


def list_input_workbooks(input_root: Path) -> dict[int, list[Path]]:
    result: dict[int, list[Path]] = {}
    for year in [2023, 2024]:
        year_dir = input_root / str(year)
        if not year_dir.exists():
            result[year] = []
            continue
        files = []
        for path in sorted(year_dir.glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            files.append(path)
        result[year] = files
    return result


def read_input_workbook(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, dtype=object, engine="openpyxl")
    return df


def sanitize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in BASE_INPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df


def template_workbooks(input_root: Path) -> list[Path]:
    report_dir = input_root / "Format_Report"
    if not report_dir.exists():
        return []
    return sorted([p for p in report_dir.glob("*.xlsx") if not p.name.startswith("~$")])


def extract_used_range(ws) -> str | None:
    rows = [r for r in ws.iter_rows() if any(c.value is not None and str(c.value).strip() != "" for c in r)]
    if not rows:
        return None
    min_row = min(cell.row for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    max_row = max(cell.row for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    min_col = min(cell.column for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    max_col = max(cell.column for row in rows for cell in row if cell.value is not None and str(cell.value).strip() != "")
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def extract_report_title(ws) -> str | None:
    title_keywords = ("TỔNG HỢP", "PHÂN TÍCH", "DANH SÁCH", "TONG HOP", "PHAN TICH", "DANH SACH")
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        values = [v for v in values if v]
        if not values:
            continue
        row_text = " ".join(values)
        upper_text = strip_accents(row_text).upper()
        if any(key in upper_text for key in title_keywords):
            if "UBND" in upper_text or "CỘNG HOÀ" in upper_text or "CỘNG HÒA" in upper_text:
                continue
            return re.sub(r"\s+", " ", row_text).strip()
    return None


def find_title_row(ws) -> tuple[int | None, str | None]:
    title_keywords = ("TỔNG HỢP", "PHÂN TÍCH", "DANH SÁCH", "TONG HOP", "PHAN TICH", "DANH SACH")
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        values = [v for v in values if v]
        if not values:
            continue
        row_text = " ".join(values)
        upper_text = strip_accents(row_text).upper()
        if any(key in upper_text for key in title_keywords):
            if "UBND" in upper_text or "CONG HOA" in upper_text or "CỘNG HÒA" in upper_text or "CỘNG HOÀ" in upper_text:
                continue
            return row_idx, re.sub(r"\s+", " ", row_text).strip()
    return None, None


def find_first_data_row(ws, start_row: int) -> int | None:
    for row_idx in range(start_row + 1, min(ws.max_row, start_row + 30) + 1):
        a_val = clean_text(ws.cell(row_idx, 1).value)
        b_val = clean_text(ws.cell(row_idx, 2).value)
        if not a_val or not b_val:
            continue
        a_text = a_val or ""
        b_norm = normalize_key(b_val or "")
        if re.fullmatch(r"\d+", a_text) and (
            b_norm.startswith("huyen") or b_norm.startswith("thanhpho") or b_norm.startswith("xa") or b_norm.startswith("phuong") or b_norm.startswith("thitran")
        ):
            return row_idx
    return None


def extract_header_rows(ws) -> list[int] | None:
    title_row, _ = find_title_row(ws)
    if title_row is None:
        return None
    first_data_row = find_first_data_row(ws, title_row)
    if first_data_row is None:
        first_data_row = min(ws.max_row, title_row + 8)
    header_rows: list[int] = []
    for row_idx in range(title_row + 1, first_data_row):
        row_values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        row_values = [v for v in row_values if v]
        if not row_values:
            continue
        row_text = " ".join(row_values)
        upper_text = strip_accents(row_text).upper()
        if "UBND" in upper_text or "CONG HOA" in upper_text or "CỘNG HÒA" in upper_text or "CỘNG HOÀ" in upper_text:
            continue
        header_rows.append(row_idx)
    return header_rows or None


def extract_district_commune_mapping(ws) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    current_district: str | None = None
    for row_idx in range(1, ws.max_row + 1):
        a_val = clean_text(ws.cell(row_idx, 1).value)
        b_val = clean_text(ws.cell(row_idx, 2).value)
        if not a_val and not b_val:
            continue
        a_text = a_val or ""
        b_text = b_val or ""
        a_norm = normalize_no_space(a_text)
        b_norm = normalize_key(b_text)
        is_district_row = bool(re.fullmatch(r"\d+", a_text)) and (b_norm.startswith("huyen") or b_norm.startswith("thanhpho"))
        is_commune_row = bool(re.fullmatch(r"\d+\.\d+", a_text)) and (
            b_norm.startswith("xa") or b_norm.startswith("phuong") or b_norm.startswith("thitran")
        )
        if is_district_row:
            current_district = b_text
            mapping.setdefault(current_district, [])
            continue
        if current_district and is_commune_row:
            commune = b_text
            if commune not in mapping[current_district]:
                mapping[current_district].append(commune)
    return mapping


def parse_template_file(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = wb[wb.sheetnames[0]]
    title_row, report_name = find_title_row(ws)
    report_name = report_name or path.stem
    header_rows = extract_header_rows(ws)
    district_commune_mapping = extract_district_commune_mapping(ws)
    return {
        "file_name": path.name,
        "report_name": report_name,
        "sheet_name": ws.title,
        "used_range": extract_used_range(ws),
        "header_rows": header_rows[:5] if header_rows else None,
        "district_commune_mapping": district_commune_mapping,
    }


def scan_report_templates(input_root: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]], list[dict[str, Any]]]:
    template_paths = template_workbooks(input_root)
    warnings: list[dict[str, Any]] = []
    if not template_paths:
        warnings.append(
            {
                "warning_type": "missing_report_templates",
                "year": None,
                "district": None,
                "message": "Format_Report is empty or unavailable; using spec fallback for schema and commune mapping.",
                "severity": "high",
            }
        )
        return build_report_schema_summary(), DISTRICT_COMMUNE_FALLBACK.copy(), warnings

    parsed = [parse_template_file(path) for path in template_paths]
    mapping: dict[str, list[str]] = {}
    for item in parsed:
        for district, communes in item["district_commune_mapping"].items():
            mapping.setdefault(district, [])
            for commune in communes:
                if commune not in mapping[district]:
                    mapping[district].append(commune)
    summary = build_report_schema_summary(template_status="parsed_from_templates")
    for item in summary:
        for parsed_item in parsed:
            if parsed_item["file_name"].startswith(f"{item['report_id']}."):
                item["sheet_name"] = parsed_item["sheet_name"]
                item["used_range"] = parsed_item["used_range"]
                item["header_rows"] = parsed_item["header_rows"]
                item["report_name"] = parsed_item["report_name"]
                break
    # Inject any observed sheet/range metadata if the template file names align with report ids.
    return summary, mapping, warnings


def build_column_coverage(processed_df: pd.DataFrame, member_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    processed_columns = set(processed_df.columns)
    member_columns = set(member_df.columns)
    for spec in REPORT_SPECS:
        for column in spec["required_columns"]:
            exists = column in processed_columns or column in member_columns
            if column.startswith("member."):
                source = "member_file"
            elif column.startswith("children.") or column.startswith("deprivation.") or column.startswith("reason.") or column.startswith("support.") or column.startswith("transition.") or column.startswith("family.") and column not in BASE_INPUT_COLUMNS:
                source = "generated"
            else:
                source = "original"
            rows.append(
                {
                    "report_id": spec["report_id"],
                    "report_name": spec["report_name"],
                    "required_column": column,
                    "exists_in_processed_dataset": bool(exists),
                    "source_type": source,
                    "note": spec["notes"].get(column, ""),
                }
            )
    return pd.DataFrame(rows)


def validate_processed_outputs(output_root: Path, mapping: dict[str, list[str]]) -> tuple[dict[str, Any], pd.DataFrame, pd.DataFrame, pd.DataFrame, list[dict[str, Any]]]:
    file_level_rows: list[dict[str, Any]] = []
    aggregate_rows: list[dict[str, Any]] = []
    null_rows: list[dict[str, Any]] = []
    legal_rows: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    dep_cols = [
        "deprivation.employment",
        "deprivation.dependentPerson",
        "deprivation.nutrition",
        "deprivation.healthInsurance",
        "deprivation.adultEducation",
        "deprivation.childSchoolAttendance",
        "deprivation.housingQuality",
        "deprivation.housingArea",
        "deprivation.cleanWater",
        "deprivation.hygienicToilet",
        "deprivation.telecommunication",
        "deprivation.informationAccessAssets",
    ]
    reason_cols = [
        "reason.lackProductionLand",
        "reason.lackCapital",
        "reason.lackLabor",
        "reason.lackProductionTools",
        "reason.lackProductionKnowledge",
        "reason.lackLaborSkill",
        "reason.illnessOrAccident",
        "reason.other",
    ]
    support_cols = [
        "support.health",
        "support.education",
        "support.production",
        "support.credit",
        "support.housing",
        "support.other",
    ]

    for year in [2023, 2024]:
        year_dir = output_root / str(year)
        if not year_dir.exists():
            continue
        for path in sorted(year_dir.glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            df = pd.read_excel(path, sheet_name="Data", engine="openpyxl", dtype=object)
            district = path.stem
            commune_allowed = set(mapping.get(district, []))
            member_path = output_root / str(year) / "_members" / f"{district}_members.xlsx"
            member_df = pd.read_excel(member_path, sheet_name="Members", engine="openpyxl", dtype=object) if member_path.exists() else pd.DataFrame()

            def _null_counts(frame: pd.DataFrame, file_type: str) -> None:
                for column in frame.columns:
                    series = frame[column]
                    null_count = int(series.isna().sum())
                    blank_count = int(series.map(lambda x: isinstance(x, str) and x.strip() == "").sum())
                    if null_count or blank_count:
                        null_rows.append(
                            {
                                "year": year,
                                "district": district,
                                "file_type": file_type,
                                "column": column,
                                "null_count": null_count,
                                "blank_count": blank_count,
                                "total_issues": null_count + blank_count,
                                "severity": "high" if null_count + blank_count > 0 else "info",
                            }
                        )

            _null_counts(df, "household")
            if not member_df.empty:
                _null_counts(member_df, "member")

            core_dup = int(df.duplicated(subset=[c for c in CORE_COLUMNS if c in df.columns]).sum()) if all(c in df.columns for c in CORE_COLUMNS) else 0
            code_dup = int(df.duplicated(subset=["family.code"]).sum()) if "family.code" in df.columns else 0

            date_parsed = df["date"].map(parse_date_value) if "date" in df.columns else pd.Series([None] * len(df))
            date_year_bad = int(date_parsed.map(lambda d: d is None or d.year != year).sum())
            candidate_dates = set(candidate_dates_for_year(year))
            date_filled_bad = int(((df.get("processing.date_was_filled", pd.Series([False] * len(df))).fillna(False).astype(bool)) & (~df.get("date", pd.Series([""] * len(df))).astype(str).isin(candidate_dates))).sum())

            missing_commune = int(df["administrative.commune"].isna().sum()) if "administrative.commune" in df.columns else len(df)
            commune_outside = int((~df["administrative.commune"].isin(commune_allowed)).sum()) if commune_allowed and "administrative.commune" in df.columns else 0
            area_null = int(df["administrative.areaType"].isna().sum()) if "administrative.areaType" in df.columns else len(df)
            b1_null = int(df["b1Point"].isna().sum()) if "b1Point" in df.columns else len(df)
            b2_null = int(df["b2Point"].isna().sum()) if "b2Point" in df.columns else len(df)
            dep_total_null = int(df["deprivation.totalCount"].isna().sum()) if "deprivation.totalCount" in df.columns else len(df)
            family_number_null = int(df["family.numberOfMembers"].isna().sum()) if "family.numberOfMembers" in df.columns else len(df)
            reviewer_null = int(df["reviewer"].isna().sum()) if "reviewer" in df.columns else len(df)
            quick_review_bad = int((df["quickReview.result"] == "--").sum()) if "quickReview.result" in df.columns else len(df)

            expected_dep_total = df[dep_cols].fillna(False).astype(bool).sum(axis=1) if all(col in df.columns for col in dep_cols) else pd.Series([0] * len(df))
            dep_total_bad = int((df["deprivation.totalCount"].fillna(0).astype(int) != expected_dep_total.astype(int)).sum()) if "deprivation.totalCount" in df.columns else len(df)
            b2_total_bad = int((df["b2Point"].fillna(0).astype(int) != (df["deprivation.totalCount"].fillna(0).astype(int) * 10)).sum()) if "b2Point" in df.columns and "deprivation.totalCount" in df.columns else len(df)

            threshold_series = df["administrative.areaType"].map(lambda x: 175 if clean_text(x) == "urban" else 140) if "administrative.areaType" in df.columns else pd.Series([140] * len(df))
            expected_classify = pd.Series([
                resolve_classification_from_scores(parse_int(b1, 0), parse_int(dep, 0), area)
                for b1, dep, area in zip(df.get("b1Point", pd.Series([0] * len(df))), df.get("deprivation.totalCount", pd.Series([0] * len(df))), df.get("administrative.areaType", pd.Series(["rural"] * len(df))))
            ])
            final_classify = df["classify.final"] if "classify.final" in df.columns else df.get("classify", pd.Series(["Hộ không nghèo"] * len(df)))
            classification_bad = int((final_classify.astype(str) != expected_classify.astype(str)).sum())
            original_classify_bad = int((df.get("classify", pd.Series(["Hộ không nghèo"] * len(df))).astype(str) == "").sum())

            dtts_bad = int(((df.get("family.isDTTS", pd.Series([False] * len(df))).fillna(False).astype(bool) & (df.get("family.ethnicity", pd.Series([""] * len(df))).astype(str) == "Kinh")).sum()))
            kinh_bad = int(((df.get("family.ethnicity", pd.Series([""] * len(df))).astype(str) == "Kinh") & (df.get("family.isDTTS", pd.Series([False] * len(df))).fillna(False).astype(bool))).sum())
            dttc_bad = int(((df.get("family.isDTTC", pd.Series([False] * len(df))).fillna(False).astype(bool)) & (~df.get("family.isDTTS", pd.Series([False] * len(df))).fillna(False).astype(bool))).sum())

            children_total = df.get("children.totalCount", pd.Series([0] * len(df))).fillna(0).astype(int)
            family_members = df.get("family.numberOfMembers", pd.Series([0] * len(df))).fillna(0).astype(int)
            children_bad = int((children_total > family_members).sum())
            child_dep_bad = int((
                (df.get("children.lackHealthInsuranceCount", pd.Series([0] * len(df))).fillna(0).astype(int) > children_total)
                | (df.get("children.nutritionDeprivedCount", pd.Series([0] * len(df))).fillna(0).astype(int) > children_total)
                | (df.get("children.schoolAttendanceDeprivedCount", pd.Series([0] * len(df))).fillna(0).astype(int) > children_total)
            ).sum())

            reason_required_bad = int(((final_classify.isin(["Hộ nghèo", "Hộ cận nghèo"])) & (df[reason_cols].fillna(False).astype(bool).sum(axis=1) == 0)).sum()) if all(col in df.columns for col in reason_cols) else 0
            support_expected = pd.Series([False] * len(df))
            if all(col in df.columns for col in support_cols):
                support_expected = pd.Series([
                    bool(
                        (row.get("deprivation.healthInsurance") == True or parse_int(row.get("children.lackHealthInsuranceCount"), 0) > 0)
                        if "support.health" else False
                    )
                    for _, row in df.iterrows()
                ])
            support_bad = 0
            if all(col in df.columns for col in support_cols):
                expected_support = []
                for _, row in df.iterrows():
                    poor_or_near = row.get("classify.final") in {"Hộ nghèo", "Hộ cận nghèo"}
                    health = bool(row.get("deprivation.healthInsurance")) or parse_int(row.get("children.lackHealthInsuranceCount"), 0) > 0
                    education = parse_int(row.get("children.schoolAttendanceDeprivedCount"), 0) > 0
                    production = bool(row.get("reason.lackProductionLand")) or bool(row.get("reason.lackProductionTools")) or bool(row.get("reason.lackProductionKnowledge"))
                    credit = bool(row.get("reason.lackCapital"))
                    housing = bool(row.get("deprivation.housingQuality")) or bool(row.get("deprivation.housingArea"))
                    other = poor_or_near and not any([health, education, production, credit, housing])
                    expected_support.append([health if poor_or_near else False, education if poor_or_near else False, production if poor_or_near else False, credit if poor_or_near else False, housing if poor_or_near else False, other if poor_or_near else False])
                expected_support_df = pd.DataFrame(expected_support, columns=support_cols)
                support_bad = int((df[support_cols].fillna(False).astype(bool).ne(expected_support_df)).any(axis=1).sum())

            transition_bad = int((df.get("transition.endingClassify", pd.Series([""] * len(df))).astype(str) != final_classify.astype(str)).sum())

            member_count_bad = 0
            host_bad = 0
            member_total_bad = 0
            if not member_df.empty and "family.code" in member_df.columns:
                member_group_sizes = member_df.groupby("family.code").size().to_dict()
                for _, row in df.iterrows():
                    code = str(row.get("family.code"))
                    expected_size = parse_int(row.get("family.numberOfMembers"), 0)
                    actual_size = int(member_group_sizes.get(code, 0))
                    if expected_size != actual_size:
                        member_total_bad += 1
                if "family.hostName" in member_df.columns:
                    host_set = set(member_df.loc[member_df.get("member.isHost", pd.Series([False] * len(member_df))).fillna(False).astype(bool), "family.hostName"].astype(str))
                    host_bad = int((~df["family.hostName"].astype(str).isin(host_set)).sum()) if "family.hostName" in df.columns else 0
            else:
                member_total_bad = len(df)

            file_level_rows.append(
                {
                    "year": year,
                    "district": district,
                    "input_rows": len(df),
                    "rows_after_core_dedup": len(df),
                    "core_duplicates_removed": 0,
                    "duplicate_family_codes_fixed": int(df.get("processing.family_code_was_changed", pd.Series([False] * len(df))).fillna(False).astype(bool).sum()),
                    "output_rows": len(df),
                    "output_file": str(path),
                }
            )

            for commune, group in df.groupby("administrative.commune", dropna=False):
                aggregate_rows.append(
                    {
                        "year": year,
                        "district": district,
                        "commune": commune,
                        "poor_households": int((group["classify.final"] == "Hộ nghèo").sum()) if "classify.final" in group.columns else 0,
                        "near_poor_households": int((group["classify.final"] == "Hộ cận nghèo").sum()) if "classify.final" in group.columns else 0,
                        "poor_members": int(group.loc[group["classify.final"] == "Hộ nghèo", "family.numberOfMembers"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                        "near_poor_members": int(group.loc[group["classify.final"] == "Hộ cận nghèo", "family.numberOfMembers"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                        "dtts_households": int(group["family.isDTTS"].fillna(False).astype(bool).sum()) if "family.isDTTS" in group.columns else 0,
                        "dttc_households": int(group["family.isDTTC"].fillna(False).astype(bool).sum()) if "family.isDTTC" in group.columns else 0,
                        "no_labor_capacity_households": int(group["family.hasNoLaborCapacity"].fillna(False).astype(bool).sum()) if "family.hasNoLaborCapacity" in group.columns else 0,
                        "revolution_merit_policy_households": int(group["family.hasRevolutionMeritPolicy"].fillna(False).astype(bool).sum()) if "family.hasRevolutionMeritPolicy" in group.columns else 0,
                        "deprivation_total_poor": int(group.loc[group["classify.final"] == "Hộ nghèo", "deprivation.totalCount"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                        "deprivation_total_near_poor": int(group.loc[group["classify.final"] == "Hộ cận nghèo", "deprivation.totalCount"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                        "children_poor_total": int(group.loc[group["classify.final"] == "Hộ nghèo", "children.totalCount"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                        "children_near_poor_total": int(group.loc[group["classify.final"] == "Hộ cận nghèo", "children.totalCount"].fillna(0).astype(int).sum()) if "classify.final" in group.columns else 0,
                    }
                )

            legal_rows.append(
                {
                    "year": year,
                    "district": district,
                    "core_duplicates": core_dup,
                    "code_duplicates": code_dup,
                    "date_year_bad": date_year_bad,
                    "date_filled_bad": date_filled_bad,
                    "missing_commune": missing_commune,
                    "commune_outside": commune_outside,
                    "area_null": area_null,
                    "b1_null": b1_null,
                    "b2_null": b2_null,
                    "deprivation_total_null": dep_total_null,
                    "b2_total_bad": b2_total_bad,
                    "deprivation_sum_bad": dep_total_bad,
                    "classification_bad": classification_bad,
                    "dtts_bad": dtts_bad,
                    "kinh_bad": kinh_bad,
                    "dttc_bad": dttc_bad,
                    "children_bad": children_bad,
                    "child_dep_bad": child_dep_bad,
                    "member_total_bad": member_total_bad,
                    "host_bad": host_bad,
                    "reason_required_bad": reason_required_bad,
                    "support_bad": support_bad,
                    "transition_bad": transition_bad,
                    "reviewer_null": reviewer_null,
                    "quick_review_bad": quick_review_bad,
                    "family_number_null": family_number_null,
                    "severity": "high" if any(
                        [
                            core_dup,
                            code_dup,
                            date_year_bad,
                            date_filled_bad,
                            missing_commune,
                            commune_outside,
                            area_null,
                            b1_null,
                            b2_null,
                            dep_total_null,
                            b2_total_bad,
                            dep_total_bad,
                            classification_bad,
                            dtts_bad,
                            kinh_bad,
                            dttc_bad,
                            children_bad,
                            child_dep_bad,
                            member_total_bad,
                            host_bad,
                            reason_required_bad,
                            support_bad,
                            transition_bad,
                            reviewer_null,
                            quick_review_bad,
                            family_number_null,
                        ]
                    ) else "info",
                }
            )

            warnings.append(
                {
                    "warning_type": "validation_check",
                    "year": year,
                    "district": district,
                    "message": f"core_dup={core_dup}, code_dup={code_dup}, date_year_bad={date_year_bad}, date_filled_bad={date_filled_bad}, commune_outside={commune_outside}, area_null={area_null}, b1_null={b1_null}, b2_null={b2_null}, dep_total_bad={dep_total_bad}, classification_bad={classification_bad}, member_total_bad={member_total_bad}, host_bad={host_bad}, reason_required_bad={reason_required_bad}, support_bad={support_bad}, transition_bad={transition_bad}",
                    "severity": "info" if not any([core_dup, code_dup, date_year_bad, date_filled_bad, commune_outside, area_null, b1_null, b2_null, dep_total_bad, classification_bad, member_total_bad, host_bad, reason_required_bad, support_bad, transition_bad]) else "high",
                }
            )

    file_level_df = pd.DataFrame(file_level_rows)
    aggregate_df = pd.DataFrame(aggregate_rows)
    warnings_df = pd.DataFrame(warnings)
    legal_consistency_df = pd.DataFrame(legal_rows)
    null_checks_df = pd.DataFrame(null_rows)
    column_coverage_df = pd.DataFrame()
    validation_frames = {
        "file_level_df": file_level_df,
        "aggregate_df": aggregate_df,
        "warnings_df": warnings_df,
        "column_coverage_df": column_coverage_df,
        "legal_consistency_df": legal_consistency_df,
        "null_checks_df": null_checks_df,
    }
    return validation_frames, file_level_df, null_checks_df, aggregate_df, warnings
